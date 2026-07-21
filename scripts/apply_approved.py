"""
apply_approved.py — generate apply_queue.json for browser pre-fill automation.

Reads the active track's applications_tracker*.xlsx for rows with Status =
"Awaiting approval", extracts profile details from the active profile.md, and
writes apply_queue.json to the JobFinder root. The browser automation step reads
this file and pre-fills each application form using Claude in Chrome.

Active track is resolved from active_track.json (written by fetch_profile.py).

Usage: python scripts/apply_approved.py
"""
import json
import pathlib
import re
from openpyxl import load_workbook

# ── Paths ────────────────────────────────────────────────────────────────────
BASE       = pathlib.Path(__file__).parent.parent   # JobFinder root
QUEUE_FILE = BASE / "apply_queue.json"

def _load_active_track():
    """Return (tracker_path, profile_path) from active_track.json, or defaults."""
    at = BASE / "active_track.json"
    if at.exists():
        try:
            data = json.loads(at.read_text(encoding="utf-8"))
            tracker_file = data.get("tracker_file", "applications_tracker.xlsx")
            profile_file = data.get("profile_file", "profile.md")
            return BASE / tracker_file, BASE / profile_file
        except Exception:
            pass
    return BASE / "applications_tracker.xlsx", BASE / "profile.md"

TRACKER, PROFILE = _load_active_track()

# ── Column indices (0-based) in the tracker ──────────────────────────────────
COL = {
    "job_id":      0,
    "date":        1,
    "source":      2,
    "company":     3,
    "role":        4,
    "location":    5,
    "work_mode":   6,
    "url":         7,
    "score":       8,
    "rationale":   9,
    "status":      10,
    "dealbreaker": 11,
    "notes":       12,
    "resume_file": 13,
    "cover_file":  14,
}


def extract_profile(profile_text: str) -> dict:
    """Extract key application fields from profile.md using regex."""

    def find(pattern, default=""):
        m = re.search(pattern, profile_text, re.IGNORECASE)
        return m.group(1).strip() if m else default

    return {
        "name":         find(r"\*\*Full name:\*\*\s*(.+)"),
        "email":        find(r"\*\*Email \(personal\):\*\*\s*([^\s—–-]+@[^\s]+)"),
        "phone":        find(r"\*\*Phone:\*\*\s*(\+?\S+)"),
        "location":     find(r"\*\*Location:\*\*\s*(.+)"),
        "linkedin":     find(r"\*\*LinkedIn:\*\*\s*(https?://\S+)"),
        "github":       find(r"\*\*GitHub:\*\*\s*(https?://\S+)"),
        "portfolio":    find(r"\*\*Portfolio:\*\*\s*(https?://\S+)"),
        "workAuth":     find(r"\*\*Citizenship / work authorization:\*\*\s*(.+?)\."),
        "sponsorship":  "No",   # "No sponsorship required for India-based roles"
        "noticePeriod": find(r"\*\*Notice period:\*\*\s*(.+)"),
        "expectedCtc":  find(r"\*\*Expected CTC:\*\*\s*(.+)"),
        "startDate":    find(r"\*\*Earliest start date:\*\*\s*(.+)"),
        "relocation":   find(r"\*\*Open to relocation:\*\*\s*(.+?)(?:\.|;|$)"),
        "workMode":     find(r"\*\*Preferred work mode:\*\*\s*(.+)"),
    }


def resolve_file_path(raw: str) -> str:
    """Return absolute path to a resume/cover file referenced in the tracker."""
    if not raw or str(raw).strip() in ("None", ""):
        return ""
    p = pathlib.Path(str(raw).strip())
    if p.is_absolute():
        return str(p)
    # Assume relative paths are under BASE
    resolved = BASE / p
    if resolved.exists():
        return str(resolved)
    return str(p)   # return as-is and let the automation handle it


def build_queue() -> list:
    """Return list of job dicts for queued applications."""
    wb = load_workbook(TRACKER, data_only=True)
    ws = wb.active

    profile_text = PROFILE.read_text(encoding="utf-8")
    profile = extract_profile(profile_text)

    queue = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[COL["job_id"]] is None:
            continue
        status = str(row[COL["status"]] or "").strip()
        if status.lower() != "awaiting approval":
            continue

        job_id  = str(row[COL["job_id"]] or "").strip()
        company = str(row[COL["company"]] or "").strip()
        role    = str(row[COL["role"]] or "").strip()
        url     = str(row[COL["url"]] or "").strip()
        score   = row[COL["score"]]
        notes   = str(row[COL["notes"]] or "").strip()

        resume_raw = str(row[COL["resume_file"]] or "").strip()
        cover_raw  = str(row[COL["cover_file"]] or "").strip()

        # Try to auto-locate the tailored documents if the tracker cell is blank
        if not resume_raw or resume_raw == "None":
            safe = f"{company}_{role}".replace("/", "_").replace(" ", "_")
            candidate = BASE / "jobs" / safe / f"{safe}_resume.docx"
            if candidate.exists():
                resume_raw = str(candidate)

        if not cover_raw or cover_raw == "None":
            safe = f"{company}_{role}".replace("/", "_").replace(" ", "_")
            candidate = BASE / "jobs" / safe / f"{safe}_cover.docx"
            if candidate.exists():
                cover_raw = str(candidate)

        queue.append({
            "job_id":       job_id,
            "company":      company,
            "role":         role,
            "url":          url,
            "score":        score,
            "notes":        notes,
            "resume_path":  resolve_file_path(resume_raw),
            "cover_path":   resolve_file_path(cover_raw),
            "fields":       profile,   # pre-fill values from profile.md
        })

    return queue


def main():
    queue = build_queue()

    if not queue:
        print("No jobs with Status='Awaiting approval' found in tracker.")
        print(f"apply_queue.json NOT written (nothing to do).")
        return

    QUEUE_FILE.write_text(
        json.dumps(queue, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )
    print(f"apply_queue.json written — {len(queue)} job(s) queued:")
    for j in queue:
        print(f"  [{j['job_id']}] {j['company']} — {j['role']}")
        print(f"       URL:    {j['url']}")
        print(f"       Resume: {j['resume_path'] or '(not found)'}")
        print(f"       Cover:  {j['cover_path'] or '(not found)'}")


if __name__ == "__main__":
    main()
