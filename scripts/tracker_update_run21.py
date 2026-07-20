"""
Daily run (2026-07-17) - append this run's 5 sourced roles to applications_tracker_NEW.xlsx. (Run 21)

  * Resolves the tracker via pathlib (BASE = this file's parent's parent), NOT a hardcoded path.
  * Dedupes on Job URL and Company+Role (safety net).
  * Assigns new Job IDs continuing from the highest JOB-### already in the sheet.

Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent          # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Luma Financial Technologies",
        "role": "AI/ML Engineer",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/luma-financial-technologies/jobs/ai-ml-engineer-8098212452",
        "score": 90,
        "rationale": "Top pick this run (LIVE; apply before 2026-09-02). India-only remote. Near-ideal fit to profile "
                     "Sec 7/9: build AI/ML + GenAI from the ground up - GenAI, agents, ML models, prompting, OCR, "
                     "vector databases and RAG; architect an agentic framework for financial advisors; own data "
                     "models + pipelines and cloud deployment. Maps one-to-one to Finance SAHAYAK (multimodal RAG, "
                     "OCR + LlamaIndex over tables/charts/text), SAHAYAK-AI (agentic, HITL) and GCP-GraphRAG. "
                     "3+ yrs and MS/PhD preferred - met (PhD in progress + 12+ yrs). AWS SageMaker/Bedrock a plus "
                     "(not required).",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, applied-AI/agentic must-have met, no sensitive fields to apply. "
                 "AWS SageMaker/Bedrock listed as a bonus only; Ollama/on-prem experience transfers. Tailored this "
                 "run. Note: fintech/structured-products domain exposure is a 'plus', not a requirement.",
        "resume": "jobs/LumaFinancialTechnologies_AIMLEngineer/LumaFinancialTechnologies_AIMLEngineer_resume.docx",
        "cover":  "jobs/LumaFinancialTechnologies_AIMLEngineer/LumaFinancialTechnologies_AIMLEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Mactores",
        "role": "Generative AI Engineer",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/mactores/jobs/generative-ai-engineer",
        "score": 85,
        "rationale": "Second tailored pick by score. India-only remote. Strong fit to profile Sec 7: design GenAI "
                     "solutions with LLMs, build scalable RAG pipelines, prompt engineering, model fine-tuning "
                     "(LoRA), vector databases, LangChain/LlamaIndex, multi-modal AI, MLOps, production APIs and "
                     "responsible AI (evaluation, bias detection). Maps to Finance SAHAYAK, GCP-GraphRAG and "
                     "SAHAYAK-AI. 3+ yrs Python - met. Minor gap: explicit AWS Bedrock/SageMaker/Lambda deployment "
                     "and LoRA fine-tuning are lighter in profile (transferable from on-prem/Ollama + transformers).",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes on skills. CAUTION - Posting apply-before date was 2026-07-08 (EXPIRED as of today). "
                 "Tailored docs prepared for reuse, but do NOT pre-fill unless the role is reposted or a live "
                 "official Mactores careers posting is found. Surface for approval / re-source. AWS deployment + LoRA "
                 "fine-tuning are minor gaps.",
        "resume": "jobs/Mactores_GenerativeAIEngineer/Mactores_GenerativeAIEngineer_resume.docx",
        "cover":  "jobs/Mactores_GenerativeAIEngineer/Mactores_GenerativeAIEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Kognitiv Inc.",
        "role": "AI Architect - Enterprise Strategy",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/kognitiv-inc/jobs/ai-architect-enterprise-strategy",
        "score": 78,
        "rationale": "Above threshold, not top 2. India-only remote, AI/ML must-have present (enterprise GenAI "
                     "architecture/strategy). Aligns with target titles 'AI Consultant / Architect' and with "
                     "SAHAYAK-AI architecture, MCP governance and multi-agent design. Leans architect/strategy-"
                     "advisory rather than hands-on build; senior-level framing fits 12+ yrs but role emphasis is "
                     "enterprise strategy over engineering depth. Full JD could not be confirmed live (aggregator "
                     "listing redirected); scored from listing metadata + skill tags.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (pending live-JD confirmation). India-only remote; AI/architecture must-have met. "
                 "Himalayas listing redirected to the generic board on fetch - confirm the posting is live and find "
                 "the official Kognitiv careers page before any pre-fill. Above threshold; not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Altisource",
        "role": "Lead AI Solution Developer",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/altisource/jobs/lead-ai-solution-developer",
        "score": 77,
        "rationale": "Above threshold, not top 2. India-only remote, AI/ML must-have present (lead AI solution "
                     "development, GenAI). Maps to full-stack AI delivery + GenAI build strengths and lead/mentoring "
                     "experience (12+ yrs). Real-estate/mortgage-tech domain. Full JD could not be confirmed live "
                     "(aggregator listing redirected); scored from listing metadata + skill tags.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (pending live-JD confirmation). India-only remote; AI build must-have met. Himalayas "
                 "listing redirected to the generic board on fetch - confirm the posting is live and find the "
                 "official Altisource careers page before any pre-fill. Above threshold; not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Cynet Corp",
        "role": "AI Prompt Engineering Lead (Agentic AI & Hiring Automation)",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/cynet-corp/jobs/ai-prompt-engineering-lead-agentic-ai-hiring-automation-remote",
        "score": 74,
        "rationale": "Above threshold, not top 2. India-only remote, AI/ML must-have present (agentic AI + prompt "
                     "engineering). Agentic AI, prompt engineering and automation are all in profile Sec 7 (SAHAYAK-AI, "
                     "MCP, n8n). BUT the role is narrower - a prompt-engineering-lead focused on hiring automation - "
                     "and the listing is ~2 months old (staleness risk). Solid partial fit behind the two tailored "
                     "roles. Scored from listing metadata + skill tags.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes on skills (agentic AI + prompt engineering in profile). Older listing (~2 months) - "
                 "confirm still live before any pre-fill. Narrower prompt-engineering/hiring-automation scope vs "
                 "research + full-stack strength. Above threshold; not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

existing_urls, existing_pairs = set(), set()
max_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    jid = str(row[0]).strip() if row[0] else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
    company = (str(row[3]).strip().lower() if len(row) > 3 and row[3] else "")
    role    = (str(row[4]).strip().lower() if len(row) > 4 and row[4] else "")
    url     = (str(row[7]).strip().lower() if len(row) > 7 and row[7] else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

next_id = max_id + 1
appended, skipped = [], []

for job in jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        skipped.append(f'{job["company"]} - {job["role"]} (already in tracker)')
        continue

    job_id = f"JOB-{next_id:03d}"
    next_id += 1

    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    appended.append(f'{job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker updated:", WB_PATH)
for a in appended:
    print("  appended", a)
for s in skipped:
    print("  skipped ", s)
