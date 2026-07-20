"""
Append JOB-036 through JOB-040 (Run 8 — 2026-07-06) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-036",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "PlexTrac",
        "role":        "AI Research Engineer - Applied AI",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/plextrac/jobs/ai-research-engineer-applied-ai",
        "score":       88,
        "rationale":   "Near-exact keyword match: RAG, LLM, SFT/RLHF, agentic AI, and explicit Model Context "
                        "Protocol (MCP) experience map directly onto SAHAYAK-AI and GCP-GraphRAG; title mirrors "
                        "profile's 'Applied AI Research Scientist' target almost verbatim; India-only remote, "
                        "3+ yrs threshold well within 12+ yrs profile.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 88). Domain is cybersecurity (SIEM/network-telemetry data) which "
                        "profile has no direct experience with — cover letter flags this honestly rather than "
                        "overstating; underlying RAG/agentic/MCP skill set transfers directly.",
        "resume":      "jobs/PlexTrac_AIResearchEngineerAppliedAI/PlexTrac_AIResearchEngineerAppliedAI_resume.docx",
        "cover":       "jobs/PlexTrac_AIResearchEngineerAppliedAI/PlexTrac_AIResearchEngineerAppliedAI_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-037",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Xenon7",
        "role":        "Senior AI/ML Engineer-Remote",
        "location":    "India (Hyderabad Remote)",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/xenon7/jobs/senior-ai-ml-engineer-remote",
        "score":       84,
        "rationale":   "Senior consulting role building 'enterprise agentic AI solutions end-to-end' with RAG, "
                        "embeddings/vector databases, and NLP/Transformers — strong overlap with SAHAYAK-AI and "
                        "GCP-GraphRAG; explicit mentoring of junior/mid engineers matches profile's PG-DAI/PGDAC "
                        "faculty experience; 6-10 yr band well within 12+ yrs profile.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second pick this run (score 84). JD lists AWS/Azure as core cloud platforms; profile's "
                        "production cloud experience is on-premise + Docker/Kubernetes rather than deep AWS/Azure "
                        "ownership — cover letter flags this gap honestly rather than overstating.",
        "resume":      "jobs/Xenon7_SeniorAIMLEngineerRemote/Xenon7_SeniorAIMLEngineerRemote_resume.docx",
        "cover":       "jobs/Xenon7_SeniorAIMLEngineerRemote/Xenon7_SeniorAIMLEngineerRemote_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-038",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Luma Financial Technologies",
        "role":        "AI/ML Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/luma-financial-technologies/jobs/ai-ml-engineer-8098212452",
        "score":       80,
        "rationale":   "GenAI, agent-based systems, OCR + vector databases + RAG for a fintech platform mirrors "
                        "the Finance SAHAYAK (OCR + LlamaIndex) project closely; postgrad-degree preference plays "
                        "to PhD-in-progress; above threshold but not a top-2 pick this run.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 80 — above threshold but third-ranked this run, so not tailored per top-2 "
                        "discipline. Fintech domain is a departure from profile's public-sector/healthcare focus "
                        "but not a deal-breaker; could revisit if a top-2 slot opens up on a future run.",
        "resume":      "",
        "cover":       "",
        "fill":        YELLOW,
    },
    {
        "job_id":      "JOB-039",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Sun King",
        "role":        "Machine Learning Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/sun-king/jobs/machine-learning-engineer",
        "score":       55,
        "rationale":   "Genuine deep-learning ML role (speech/audio ASR/TTS + computer vision, Wav2Vec/Whisper/"
                        "YOLO) but no GenAI, RAG, or agentic-AI component; profile has no direct production "
                        "speech/vision experience at this depth — below threshold.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 55 — below threshold, not tailored. Has a genuine AI/ML component (deep learning) "
                        "so not a deal-breaker, just a weak match to profile's GenAI/RAG/agentic-AI focus.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-040",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Hypersonix",
        "role":        "Data Scientist_ML (India)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/hypersonix/jobs/data-scientist_ml-india",
        "score":       42,
        "rationale":   "Traditional ML role for retail pricing optimisation and demand forecasting — no GenAI, "
                        "RAG, agentic-AI, or multilingual-NLP component; weak alignment with profile's research "
                        "focus and target titles.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 42 — well below threshold, not tailored. Has an ML component so not a "
                        "deal-breaker, just a poor fit against profile's GenAI-research emphasis.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-036 through JOB-040 appended to {WB_PATH}")
