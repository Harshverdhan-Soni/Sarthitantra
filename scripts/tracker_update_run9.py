"""
Append JOB-041 through JOB-045 (Run 9 — 2026-07-08) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-041",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Sia Partners",
        "role":        "Staff GenAI Engineer",
        "location":    "India (Mumbai)",
        "work_mode":   "Remote/Hybrid",
        "url":         "https://himalayas.app/companies/sia-partners/jobs/staff-genai-engineer",
        "score":       82,
        "rationale":   "Near-perfect GenAI/RAG/agentic-workflow match (LLM-based systems, RAG architectures, "
                        "multi-step reasoning, tool-calling, memory management, LangChain/Google ADK) mapping "
                        "directly onto SAHAYAK-AI and GCP-GraphRAG; India-based (Mumbai); 8+ yr bar comfortably "
                        "met by 12+ yrs profile.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 82). Gap: JD leans heavily on Terraform/IaC and cloud-native "
                        "microservices-at-scale ownership vs profile's Docker/Kubernetes + on-premise depth — "
                        "cover letter flags this honestly rather than overstating cloud ownership.",
        "resume":      "jobs/SiaPartners_StaffGenAIEngineer/SiaPartners_StaffGenAIEngineer_resume.docx",
        "cover":       "jobs/SiaPartners_StaffGenAIEngineer/SiaPartners_StaffGenAIEngineer_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-042",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "HighLevel",
        "role":        "Software Development Engineer III - Gen AI (Workflow AI/RAG)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/highlevel/jobs/software-development-engineer-iii-gen-ai-7901534214",
        "score":       76,
        "rationale":   "RAG systems + knowledge-base + context-aware workflow engineering directly matches "
                        "GCP-GraphRAG; mentoring-junior-engineers requirement matches PG-DAI/PGDAC faculty "
                        "experience; India remote, large stable employer (1001-5000 employees).",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second pick this run (score 76). Gap: role expects low-latency engineering at "
                        "HighLevel's extreme scale (billions of API hits/day, 250+ microservices) vs profile's "
                        "national-government-scale API delivery — cover letter names this gap directly rather "
                        "than overstating scale experience.",
        "resume":      "jobs/HighLevel_SDE3-GenAI-WorkflowRAG/HighLevel_SDE3-GenAI-WorkflowRAG_resume.docx",
        "cover":       "jobs/HighLevel_SDE3-GenAI-WorkflowRAG/HighLevel_SDE3-GenAI-WorkflowRAG_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-043",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Nucs AI",
        "role":        "Machine Learning Scientist (Clinical Oncology)",
        "location":    "Remote (worldwide, incl. India)",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/nucs-ai/jobs/machine-learning-scientist",
        "score":       38,
        "rationale":   "Genuine ML/DL research role but requires a PhD specifically in medical imaging / "
                        "computer vision with a publication record in clinical-imaging venues (PET/CT, MRI, "
                        "PSMA-PET, SPECT, nuclear-medicine dosimetry); profile's GenAI/RAG/NLP research and "
                        "HMIS healthcare-delivery background do not transfer to this deep clinical-imaging "
                        "specialisation.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 38 — well below threshold, not tailored. Has a genuine AI/ML component (deep "
                        "learning research) so not a deal-breaker, just a hard domain-specialisation mismatch.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-044",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Miratech",
        "role":        "Researcher (Recruiting / Candidate Sourcing)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/miratech/jobs/researcher-8660438504",
        "score":       None,
        "rationale":   "Role is a talent-sourcing / recruiting researcher (active candidate search via "
                        "LinkedIn, professional communities, pipeline building) with no AI/ML or technical "
                        "engineering component whatsoever.",
        "status":      "Skipped",
        "dealbreaker": "Yes",
        "notes":       "DEAL-BREAKER: No AI/ML component — this is an HR/recruiting-sourcing role, not an "
                        "engineering or research position. Listing title 'Researcher' was misleading relative "
                        "to actual JD content.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-045",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Yaqeen Institute",
        "role":        "Research Scientist (Global Muslim Studies)",
        "location":    "Remote (Global)",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/yaqeen-institute/jobs/research-scientist",
        "score":       None,
        "rationale":   "Social-science / religious-studies research role designing and analysing empirical "
                        "studies on the religious, psychological, and sociological dynamics of Muslim "
                        "populations globally — no AI/ML, engineering, or technical component.",
        "status":      "Skipped",
        "dealbreaker": "Yes",
        "notes":       "DEAL-BREAKER: No AI/ML component — pure social-science research role at a nonprofit "
                        "Islamic research institute, unrelated to profile's AI/ML target titles.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-041 through JOB-045 appended to {WB_PATH}")
