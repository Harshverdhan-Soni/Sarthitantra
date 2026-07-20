"""
Run 12 (2026-07-10) - append this run's sourced roles to applications_tracker_NEW.xlsx.

Robust behaviour:
  * Resolves the tracker via pathlib (BASE = this file's parent's parent), NOT a hardcoded path.
  * Reads existing rows and SKIPS any role whose Job URL, or Company+Role, already exists
    (dedupe safety net in case a role was logged in a prior run).
  * Assigns new Job IDs continuing from the highest JOB-### already in the sheet.

Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent          # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns A-S:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode | Job URL |
# Fit score | Score rationale | Status | Deal-breaker? | Missing requirements / notes |
# Resume file | Cover letter file | Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Proximity Works",
        "role": "GenAI Data Scientist (Search, Discovery)",
        "location": "India", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/proximity-works/jobs/genai-data-scientist-search-discovery",
        "score": 85,
        "rationale": "Top pick this run. GenAI + search/discovery maps directly onto candidate's core "
                     "retrieval work: embedding-based semantic search, Neo4j knowledge-graph ranking, and "
                     "GNN-driven context pruning (GCP-GraphRAG, IEEE GCON 2026), plus multimodal document "
                     "retrieval (Finance SAHAYAK). India-only remote; strong Python/PyTorch fit.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Full JD text could not be fetched (job detail page is client-rendered and "
                 "Claude in Chrome is disallowed in the scheduled run) - scored from listing title, seniority, "
                 "India-only tag and company profile; verify exact JD requirements before submitting.",
        "resume": "jobs/ProximityWorks_GenAIDataScientist-SearchDiscovery/ProximityWorks_GenAIDataScientist-SearchDiscovery_resume.docx",
        "cover":  "jobs/ProximityWorks_GenAIDataScientist-SearchDiscovery/ProximityWorks_GenAIDataScientist-SearchDiscovery_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "SS&C Technologies",
        "role": "GenAI Developer / Engineer",
        "location": "India", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/ss-c-technologies/jobs/genai-developer-engineer-2874542653",
        "score": 84,
        "rationale": "Second pick. GenAI developer/engineer in financial + healthcare tech aligns with "
                     "candidate's regulated-domain GenAI work: Finance SAHAYAK (multimodal RAG over financial "
                     "documents), SAHAYAK-AI agents with governance, plus strong Java/Spring Boot + Python "
                     "delivery. India-only remote; freshest posting this run.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. JD text not fetchable (client-rendered page; Chrome disallowed) - scored from "
                 "listing metadata and company profile. Confirm required cloud stack (Azure/AWS Bedrock vs "
                 "on-prem Ollama) before submitting.",
        "resume": "jobs/SSCTechnologies_GenAIDeveloperEngineer/SSCTechnologies_GenAIDeveloperEngineer_resume.docx",
        "cover":  "jobs/SSCTechnologies_GenAIDeveloperEngineer/SSCTechnologies_GenAIDeveloperEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Nagarro",
        "role": "Associate Staff Engineer, Generative AI",
        "location": "India", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/nagarro/jobs/associate-staff-engineer-generative-ai",
        "score": 82,
        "rationale": "Strong senior GenAI engineering match (LLMs, RAG, agentic AI, Python, cloud, scalable "
                     "services). Above threshold but not top-2 this run; distinct role/URL from prior Nagarro "
                     "entries in the tracker.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Above threshold, not tailored (only top 2 tailored per run). JD text not "
                 "fetchable in scheduled run; scored from listing metadata.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "PradeepIT Consulting Services",
        "role": "Tech Lead - Gen AI (6+ Years)",
        "location": "India", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/pradeepit-consulting-services-pvt-ltd/jobs/tech-lead-gen-ai-with-6-years-1590182615",
        "score": 76,
        "rationale": "GenAI tech-lead role; 6+ yr bar comfortably met by 12+ yr profile; LLM/RAG/agent scope "
                     "aligns with SAHAYAK-AI and GCP-GraphRAG. Above threshold but not top-2 this run.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Above threshold, not tailored. Consulting-firm role (staffing-oriented); "
                 "JD text not fetchable in scheduled run - scored from listing metadata.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Zensar",
        "role": "Generative AI Application Developer",
        "location": "India", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/zensar/jobs/aes-de-generative-ai-application-developers",
        "score": 74,
        "rationale": "GenAI application-developer role with LLM/RAG build scope matching profile; India-only "
                     "remote. Above threshold but not top-2; enterprise-services delivery context.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Above threshold, not tailored. JD text not fetchable in scheduled run - scored "
                 "from listing metadata; confirm seniority level (appears mid-level) before applying.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

# --- build dedupe sets and find the highest existing JOB-### ---
existing_urls, existing_pairs = set(), set()
max_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    jid = str(row[0]).strip() if row[0] else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
    company = (str(row[3]).strip().lower() if len(row) > 3 and row[3] else "")
    role    = (str(row[4]).strip().lower() if len(row) > 4 and row[4] else "")
    url     = (str(row[7]).strip().lower() if len(row) > 7 and row[7] else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

next_id = max_id + 1
appended, skipped = [], []

for job in jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        skipped.append(f'{job["company"]} - {job["role"]} (already in tracker)')
        continue

    job_id = f"JOB-{next_id:03d}"
    next_id += 1

    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    appended.append(f'{job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker updated:", WB_PATH)
for a in appended:
    print("  appended", a)
for s in skipped:
    print("  skipped ", s)
