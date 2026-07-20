"""
Append JOB-046 through JOB-050 (Run 10 — 2026-07-08) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-046",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Srijan Technologies",
        "role":        "Lead AI Engineer (Lead Agentic AI Engineer)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/srijan-technologies/jobs/lead-ai-engineer",
        "score":       87,
        "rationale":   "Near-perfect agentic AI match: multi-agent architecture (orchestrator/sub-agent "
                        "patterns), advanced RAG, guardrails, and MLOps/production deployment map directly "
                        "onto SAHAYAK-AI and GCP-GraphRAG; India-only remote; 5-10 yr bar comfortably met "
                        "by 12+ yrs profile.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 87). Gap: JD names specific frameworks (LangGraph, CrewAI, "
                        "AutoGen, Semantic Kernel) and production LLM-serving tooling (vLLM, DeepSpeed, "
                        "Triton, LangSmith/Arize) not explicit in profile — SAHAYAK-AI's MCP-based "
                        "orchestration covers the same architecture patterns; cover letter names this gap "
                        "honestly. Client-services company (Material/Srijan), not an in-house product team.",
        "resume":      "jobs/Srijan_LeadAIEngineer/Srijan_LeadAIEngineer_resume.docx",
        "cover":       "jobs/Srijan_LeadAIEngineer/Srijan_LeadAIEngineer_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-047",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "QuantumLoopAI",
        "role":        "AI/ML & Prompt Engineer — LLM, RAG & Voice Agent",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/quantumloopai/jobs/ai-ml-prompt-engineer-llm-rag-voice-agent",
        "score":       83,
        "rationale":   "Core LLM/RAG/prompt-engineering role for a healthcare-AI product (NHS GP-surgery AI "
                        "receptionist) closely parallels HMIS national healthcare-delivery experience and "
                        "GCP-GraphRAG's retrieval/prompt-tuning work; India remote, senior level, no degree "
                        "required.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second pick this run (score 83). Gap: role is built on NestJS/Node.js plus voice-"
                        "agent/STT-TTS/telephony integration, none of which are in profile (Python/Java "
                        "background, no voice-agent shipped work) — cover letter flags this honestly. UK "
                        "business-hours (9am-6pm GMT/BST) alignment required.",
        "resume":      "jobs/QuantumLoopAI_AIMLPromptEngineer/QuantumLoopAI_AIMLPromptEngineer_resume.docx",
        "cover":       "jobs/QuantumLoopAI_AIMLPromptEngineer/QuantumLoopAI_AIMLPromptEngineer_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-048",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Precision Medicine Group (Precision AQ)",
        "role":        "AI/ML Engineer II",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/precision-medicine-group/jobs/ai-ml-engineer-ii",
        "score":       80,
        "rationale":   "GenAI/LLM fine-tuning and evaluation work for an oncology-access/healthcare-analytics "
                        "product aligns strongly with HMIS national healthcare-delivery background; India "
                        "remote, 3+ yr bar comfortably met by 12+ yrs profile.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 80 — above threshold but third-ranked this run, so not tailored per top-2 "
                        "discipline. Strong backup candidate if either top-2 pick falls through; healthcare/"
                        "oncology domain is a good match to HMIS experience.",
        "resume":      "",
        "cover":       "",
        "fill":        YELLOW,
    },
    {
        "job_id":      "JOB-049",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Tech Holding",
        "role":        "ML / AI Data Engineer (Contract)",
        "location":    "India",
        "work_mode":   "Remote (Contract/Temporary)",
        "url":         "https://himalayas.app/companies/tech-holding/jobs/ml-ai-data-engineer-contract",
        "score":       56,
        "rationale":   "Senior data-pipeline / GPU-distributed-systems role for large-scale video/audio ML "
                        "workflows (Spark, Ray, Kafka, Airflow); no GenAI, LLM, RAG, or agentic-AI component "
                        "in the JD — pure ML-infrastructure/data-engineering fit rather than profile's GenAI "
                        "specialisation.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 56 — below 70 threshold, not tailored. Has a genuine ML-adjacent component so "
                        "not a deal-breaker. Flag: listing's 'Apply before' date is 2026-07-08 (today) — "
                        "may already be closed by the time this is reviewed. Contract/temporary, not "
                        "full-time.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-050",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Bizmoni Corp.",
        "role":        "Machine Learning (ML) Ops Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/bizmoni-corp/jobs/machine-learning-ml-ops-engineer",
        "score":       32,
        "rationale":   "Genuine MLOps role (Docker/Kubernetes/CI-CD/Prometheus-Grafana pipeline automation) "
                        "but listing explicitly states compensation is ONLY equity-based for a ~16 hrs/week "
                        "commitment despite being tagged 'Full Time' — fundamentally misaligned with "
                        "profile's need for a salaried position (current CTC 7LPA, expected 15LPA).",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 32 — well below threshold, not tailored. Not treated as a literal deal-breaker "
                        "(has an AI/ML component per profile Section 9 rule) but flagged here as effectively "
                        "unsuitable: equity-only compensation, part-time hours (16 hrs/week), early-stage "
                        "startup with no salary disclosed. Recommend skipping regardless of any future score "
                        "recalculation unless compensation terms change.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-046 through JOB-050 appended to {WB_PATH}")
