"""
pending_confirmations.py - list applications waiting for the user's own Submit.

Claude can never click Submit, so roles sit in the tracker until the user applies
and runs mark_applied.py. Groups what is outstanding:
  1. Pre-filled, ready to submit  (Status = "Awaiting approval", no Date submitted)
  2. Tailored, not yet submitted  (Status = "Tailored", no Date submitted)

Usage:
    python pending_confirmations.py          # recent tailored shown = 6
    python pending_confirmations.py --all    # list every tailored role too
"""
import sys
import pathlib
from openpyxl import load_workbook

BASE = pathlib.Path(__file__).parent.parent
WB_PATH = str(BASE / "applications_tracker_NEW.xlsx")

COL_ID = 1
COL_COMPANY = 4
COL_ROLE = 5
COL_STATUS = 11
COL_DATE_SUB = 16

SHOW_ALL = "--all" in sys.argv[1:]
RECENT_N = 6
BAR = "=" * 70


def collect():
    wb = load_workbook(WB_PATH, read_only=True)
    ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
    awaiting, tailored = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        if row[COL_DATE_SUB - 1]:
            continue
        status = str(row[COL_STATUS - 1] or "").strip().lower()
        entry = (row[COL_ID - 1], row[COL_COMPANY - 1], row[COL_ROLE - 1])
        if status == "awaiting approval":
            awaiting.append(entry)
        elif status == "tailored":
            tailored.append(entry)
    return awaiting, tailored


def main():
    awaiting, tailored = collect()
    if not awaiting and not tailored:
        print("No applications awaiting your confirmation.")
        return

    print(BAR)
    print("SUBMISSION CONFIRMATION - Claude cannot click Submit for you.")
    print("Tell me which of these you have applied to and I will mark them Submitted.")
    print(BAR)

    if awaiting:
        print("")
        print(">> PRE-FILLED and READY FOR YOU TO SUBMIT:")
        for jid, company, role in awaiting:
            print("     [" + str(jid) + "] " + str(company) + " - " + str(role))

    if tailored:
        shown = tailored if SHOW_ALL else tailored[-RECENT_N:]
        tail = "" if SHOW_ALL else (", showing latest " + str(len(shown)))
        print("")
        print(">> TAILORED, NOT YET SUBMITTED (" + str(len(tailored)) + " total" + tail + "):")
        for jid, company, role in shown:
            print("     [" + str(jid) + "] " + str(company) + " - " + str(role))
        extra = len(tailored) - len(shown)
        if not SHOW_ALL and extra > 0:
            print("     ... and " + str(extra) + " more - run with --all to list them.")

    print("")
    print("-" * 70)
    print("Applied to any yourself? Record it so the tracker stays accurate:")
    print("    python mark_applied.py JOB-XXX   (or double-click mark_applied.bat)")
    print(BAR)


if __name__ == "__main__":
    main()
