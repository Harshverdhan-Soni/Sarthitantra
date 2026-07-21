"""
mark_applied.py - record that YOU (the user) have successfully submitted an application.

Claude can pre-fill applications but can never click Submit, so once you finish an
application yourself, run this to mark it Submitted in applications_tracker_NEW.xlsx.

Usage (from the scripts/ folder):
    python mark_applied.py                     -> interactive: it asks for the Job ID
    python mark_applied.py JOB-014             -> mark one job submitted (date = today)
    python mark_applied.py JOB-014 2026-07-16  -> mark submitted on a specific date
    python mark_applied.py JOB-014 --outcome "Applied via BuzzBoard ATS"

What it does for the given Job ID:
    * Status            -> "Submitted"
    * Date submitted    -> the date you applied (today unless you pass one)
    * Follow-up date    -> +7 days (only if empty)
    * Outcome           -> optional note you pass with --outcome (appended)
    * Last updated      -> today
    * Row colour        -> blue (#CFE2F3) so submitted rows stand out
It never deletes anything and only touches the row(s) you name.
"""
import sys, re, pathlib
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

import json as _json

BASE = pathlib.Path(__file__).parent.parent

def _tracker_path():
    at = BASE / "active_track.json"
    if at.exists():
        try:
            data = _json.loads(at.read_text(encoding="utf-8"))
            return str(BASE / data.get("tracker_file", "applications_tracker.xlsx"))
        except Exception:
            pass
    return str(BASE / "applications_tracker.xlsx")

WB_PATH = _tracker_path()

BLUE = PatternFill("solid", fgColor="CFE2F3")   # submitted
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

COL_ID, COL_COMPANY, COL_ROLE = 1, 4, 5
COL_STATUS, COL_DATE_SUB, COL_FOLLOWUP, COL_OUTCOME, COL_UPDATED = 11, 16, 17, 18, 19


def parse_args(argv):
    job_ids, submit_date, outcome = [], date.today().isoformat(), None
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--outcome":
            outcome = argv[i + 1] if i + 1 < len(argv) else None
            i += 2
            continue
        if re.match(r"^\d{4}-\d{2}-\d{2}$", a):
            submit_date = a
        elif re.match(r"(?i)^job-?\d+$", a):
            m = re.match(r"(?i)job-?(\d+)", a)
            job_ids.append(f"JOB-{int(m.group(1)):03d}")
        i += 1
    return job_ids, submit_date, outcome


def mark(ws, job_id, submit_date, outcome):
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, COL_ID).value).strip().upper() == job_id:
            ws.cell(r, COL_STATUS).value = "Submitted"
            ws.cell(r, COL_DATE_SUB).value = submit_date
            if not (ws.cell(r, COL_FOLLOWUP).value or "").strip() if isinstance(ws.cell(r, COL_FOLLOWUP).value, str) else not ws.cell(r, COL_FOLLOWUP).value:
                ws.cell(r, COL_FOLLOWUP).value = (date.fromisoformat(submit_date) + timedelta(days=7)).isoformat()
            if outcome:
                cur = ws.cell(r, COL_OUTCOME).value or ""
                note = f"Submitted by user on {submit_date}. {outcome}".strip()
                ws.cell(r, COL_OUTCOME).value = (cur + " | " + note) if cur else note
            ws.cell(r, COL_UPDATED).value = date.today().isoformat()
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = BLUE
                ws.cell(r, c).alignment = WRAP
            ws.cell(r, COL_ID).font = BOLD
            return r, ws.cell(r, COL_COMPANY).value, ws.cell(r, COL_ROLE).value
    return None, None, None


def main():
    job_ids, submit_date, outcome = parse_args(sys.argv[1:])
    if not job_ids:
        try:
            entered = input("Enter the Job ID you applied to (e.g. JOB-014): ").strip()
        except EOFError:
            entered = ""
        m = re.match(r"(?i)job-?(\d+)", entered)
        if not m:
            print("No valid Job ID given. Nothing changed.")
            return
        job_ids = [f"JOB-{int(m.group(1)):03d}"]

    wb = load_workbook(WB_PATH)
    ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active
    changed = []
    for jid in job_ids:
        row, company, role = mark(ws, jid, submit_date, outcome)
        if row:
            changed.append(f"{jid}: {company} - {role}  ->  Submitted ({submit_date})")
        else:
            changed.append(f"{jid}: NOT FOUND in tracker")
    wb.save(WB_PATH)
    print("Tracker updated:", WB_PATH)
    for c in changed:
        print("  ", c)


if __name__ == "__main__":
    main()
