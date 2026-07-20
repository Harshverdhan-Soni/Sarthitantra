"""
Daily run (2026-07-15) - append this run's 5 sourced roles to applications_tracker_NEW.xlsx. (Run 16)

  * Resolves the tracker via pathlib (BASE = this file's parent's parent), NOT a hardcoded path.
  * Dedupes on Job URL and Company+Role (safety net).
  * Assigns new Job IDs continuing from the highest JOB-### already in the sheet.

Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent          # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Srijan Technologies (Material)",
        "role": "AI Engineer - Agentic",
        "location": "India (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/srijan-technologies/jobs/ai-engineer-agentic",
        "score": 90,
        "rationale": "Top pick this run. India-only remote. Near-perfect map to profile Sec 7: build multi-step agentic "
                     "workflows with LangGraph/LangChain/AutoGen/CrewAI/LlamaIndex, RAG pipelines (chunking, embeddings, "
                     "vector search, re-ranking), tool-calling, memory, human-in-the-loop, FastAPI, Docker/K8s, CI/CD, "
                     "observability. Directly matches SAHAYAK-AI (multi-agent, MCP, RBAC, HITL) and GCP-GraphRAG "
                     "(RAG + Neo4j KG, GCON 2026). 2-5 yrs asked vs 12+ yrs - comfortably clears.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, GenAI must-have met, no sensitive fields to apply. MLOps/K8s "
                 "productionization is 'non-negotiable' - profile shows Docker/K8s/CI-CD so covered. Apply-before "
                 "Sep 03 2026. Consulting delivery (Material/Srijan).",
        "resume": "jobs/Srijan_AIEngineer-Agentic/Srijan_AIEngineer-Agentic_resume.docx",
        "cover":  "jobs/Srijan_AIEngineer-Agentic/Srijan_AIEngineer-Agentic_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Paralucent",
        "role": "GenAI Full Stack Developer - India (PL854)",
        "location": "India (Remote)", "work_mode": "Remote (3-month contract)",
        "url": "https://himalayas.app/companies/paralucent/jobs/genai-full-stack-developer-india-pl854",
        "score": 87,
        "rationale": "Second pick. India-only remote. Rare combo that hits both of Harsh's strengths: enterprise RAG "
                     "(ingestion, chunking, embeddings, vector/hybrid search, re-ranking, grounding, hallucination "
                     "mitigation) AND full-stack (React/Next.js + Python/FastAPI + Java/.NET, OAuth2/JWT/RBAC, REST/"
                     "microservices). Maps to GCP-GraphRAG + national-scale Java/Spring Boot/ReactJS delivery. "
                     "3-8+ yrs - clears. Azure OpenAI depth is lighter (upskill flag, not a block).",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, GenAI must-have met. Contract (3 months) - confirm acceptable "
                 "before applying. Azure-native stack (Azure OpenAI/AI Search/Functions) preferred; profile has "
                 "Azure familiarity but Azure OpenAI depth lighter - flag as upskill area. Apply-before Jul 30 2026.",
        "resume": "jobs/Paralucent_GenAIFullStackDeveloper/Paralucent_GenAIFullStackDeveloper_resume.docx",
        "cover":  "jobs/Paralucent_GenAIFullStackDeveloper/Paralucent_GenAIFullStackDeveloper_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "PradeepIT Consulting Services Pvt Ltd",
        "role": "Sr. Engineer - Gen AI with 4+ years",
        "location": "India (Remote)", "work_mode": "Remote (6-month extendable)",
        "url": "https://himalayas.app/companies/pradeepit-consulting-services-pvt-ltd/jobs/sr-engineer-gen-ai-with-4-years",
        "score": 76,
        "rationale": "Above threshold, not tailored (outside top 2). India-only remote GenAI backend role, 4+ yrs. "
                     "Good GenAI/Python/backend/cloud fit with profile Sec 7 (text generation, Python, AWS/Azure, "
                     "SQL/NoSQL). Below the tailored two: scope is backend-integration of GenAI models rather than "
                     "agentic/RAG research depth, and CV/audio-gen emphasis is secondary to Harsh's NLP/RAG core. "
                     "Distinct from already-logged PradeepIT 'Tech Lead - Gen AI (6+ Years)' (JOB-034).",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, GenAI must-have met. Contract/backend delivery scope; confirm full "
                 "JD before applying. Not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "micro1",
        "role": "Machine Learning Engineer",
        "location": "Remote (all countries, incl. India)", "work_mode": "Remote (Contractor)",
        "url": "https://himalayas.app/companies/micro1/jobs/machine-learning-engineer-7461269554",
        "score": 62,
        "rationale": "Below threshold. Remote incl. India, mid-level contractor. Classical ML stack (TensorFlow, "
                     "scikit-learn, data preprocessing, feature engineering) with 'no prior AI experience required' - "
                     "no GenAI/RAG/agentic/LLM component, which is Harsh's core strength per Sec 7/9. AI/ML present so "
                     "not a hard deal-breaker, but mid-level classical-ML scope is well below a senior/12+-yr, "
                     "PhD-in-progress GenAI profile. Not tailored.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (classical-ML, mid-level scope vs senior GenAI/research target in profile Sec 9). "
                 "AI/ML component present so not a strict deal-breaker. Below threshold - not tailored. "
                 "Apply-before Jul 25 2026.",
        "resume": "", "cover": "", "fill": PINK,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "INDG (Grip)",
        "role": "Senior AI Engineer",
        "location": "India (Remote, multi-country)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/indg/jobs/senior-ai-engineer",
        "score": 55,
        "rationale": "Below threshold. India among eligible countries, senior. But domain is diffusion-model / ComfyUI "
                     "visual-content generation (image pipelines, latent-space control, Blender/Photoshop) - a computer-"
                     "vision/creative-AI focus that does not match Harsh's NLP/RAG/agentic core (Sec 7). AI/ML present so "
                     "not a hard deal-breaker, but poor domain fit; also apply-by date has passed (Jul 08 2026). "
                     "Not tailored.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (diffusion/ComfyUI computer-vision domain vs NLP/RAG/agentic strengths in profile Sec 7). "
                 "AI/ML component present so not a strict deal-breaker. Listing appears EXPIRED (apply-before Jul 08 "
                 "2026). Below threshold - not tailored.",
        "resume": "", "cover": "", "fill": PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

existing_urls, existing_pairs = set(), set()
max_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    jid = str(row[0]).strip() if row[0] else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
    company = (str(row[3]).strip().lower() if len(row) > 3 and row[3] else "")
    role    = (str(row[4]).strip().lower() if len(row) > 4 and row[4] else "")
    url     = (str(row[7]).strip().lower() if len(row) > 7 and row[7] else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

next_id = max_id + 1
appended, skipped = [], []

for job in jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        skipped.append(f'{job["company"]} - {job["role"]} (already in tracker)')
        continue

    job_id = f"JOB-{next_id:03d}"
    next_id += 1

    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    appended.append(f'{job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker updated:", WB_PATH)
for a in appended:
    print("  appended", a)
for s in skipped:
    print("  skipped ", s)
