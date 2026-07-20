"""
Append JOB-031 through JOB-035 (Run 7 — 2026-07-05) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-031",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Proximity Works",
        "role":        "GenAI - Data Scientist (Search, Discovery)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/proximity-works/jobs/genai-data-scientist-search-discovery",
        "score":       70,
        "rationale":   "GenAI application to search/discovery aligns with RAG/vector-search/embeddings skill set; "
                        "title is Data Scientist rather than a research/engineer title, and boutique dev-studio "
                        "context (vs. research or public-sector focus) caps the score at threshold.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score exactly at threshold (70) but not one of the top 2 this run, so not tailored per "
                        "daily-cap discipline. Could revisit if a top-2 slot opens up.",
        "resume":      "",
        "cover":       "",
        "fill":        YELLOW,
    },
    {
        "job_id":      "JOB-032",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Bridge-it",
        "role":        "Fractional AI Architect (Consultant)",
        "location":    "India",
        "work_mode":   "Remote (Contract/Fractional, 1-3 months)",
        "url":         "https://himalayas.app/companies/bridge-it/jobs/fractional-ai-architect-consultant",
        "score":       84,
        "rationale":   "Near-perfect keyword match: Generative AI, RAG, LLM, vector databases, knowledge graphs, "
                        "agent orchestration, and prompt engineering map directly onto GCP-GraphRAG and SAHAYAK-AI; "
                        "12+ yrs minimum bar exactly matches profile; title 'AI Consultant' is an explicit target "
                        "title in profile Section 9.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 84). Flags: (1) this is a fractional/contract engagement "
                        "(1-3 months), not a full-time role — confirm fit with notice-period/CTC expectations "
                        "before proceeding; (2) JD prefers hands-on AWS/GCP/Azure production ownership, whereas "
                        "profile's infra experience is on-premise + Docker/Kubernetes — cover letter flags this "
                        "honestly rather than overstating cloud experience; (3) EdTech/US K-12 domain experience "
                        "is 'highly desirable' but not held — not treated as a deal-breaker since it's not a "
                        "must-have.",
        "resume":      "jobs/Bridge-it_FractionalAIArchitect-Consultant/Bridge-it_FractionalAIArchitect-Consultant_resume.docx",
        "cover":       "jobs/Bridge-it_FractionalAIArchitect-Consultant/Bridge-it_FractionalAIArchitect-Consultant_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-033",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Wadhwani AI",
        "role":        "Machine Learning Scientist (LLM)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/wadhwani-ai/jobs/machine-learning-scientist-llm-5076052172",
        "score":       80,
        "rationale":   "LLM-focused ML Scientist role at a nonprofit applying AI to underserved-population "
                        "problems (health, agriculture, public good) — strong match to profile's public-sector/"
                        "healthcare-AI nice-to-have and multilingual-NLP/RAG research background (GCP-GraphRAG, "
                        "GCON 2026); distinct role from the previously tailored Wadhwani AI ASR posting (JOB-019).",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second pick this run (score 80). Full JD text could not be retrieved this run (page did "
                        "not render on repeated fetch attempts); scoring and tailoring are based on the listed "
                        "title, company mission, and skill tags visible in search results — recommend the user "
                        "double-check the live JD for any additional hard requirements before applying.",
        "resume":      "jobs/WadhwaniAI_MachineLearningScientist-LLM/WadhwaniAI_MachineLearningScientist-LLM_resume.docx",
        "cover":       "jobs/WadhwaniAI_MachineLearningScientist-LLM/WadhwaniAI_MachineLearningScientist-LLM_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-034",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "PradeepIT Consulting Services Pvt Ltd",
        "role":        "Tech Lead - Gen AI (6+ Years)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/pradeepit-consulting-services-pvt-ltd/jobs/tech-lead-gen-ai-with-6-years-1590182615",
        "score":       64,
        "rationale":   "Genuine GenAI focus but at an IT-staffing/consulting firm placing candidates onto client "
                        "projects — title (Tech Lead) and organisational context are a weaker match than profile's "
                        "research/architecture-level target titles; below the 70 threshold.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 64 — below threshold, not tailored. Has a genuine GenAI component so not a "
                        "deal-breaker, just a weaker organisational/seniority fit.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-035",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Zensar",
        "role":        "AES - DE - Generative AI Application Developers",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/zensar/jobs/aes-de-generative-ai-application-developers",
        "score":       58,
        "rationale":   "Generative AI focus present but the listing reads as an internal grade-coded (AES/DE) "
                        "developer requisition at a large IT-services firm — implies project-delivery/application "
                        "development rather than research or architecture ownership; below threshold.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 58 — below threshold, not tailored. Has an AI/ML component so not a deal-breaker, "
                        "just a weaker seniority/research fit than profile target.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-031 through JOB-035 appended to {WB_PATH}")
