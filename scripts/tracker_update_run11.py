"""
Append JOB-051 through JOB-055 (Run 11 — 2026-07-09) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-051",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "MyAdvice",
        "role":        "AI Engineer I (Generative AI)",
        "location":    "India (Bangalore preferred)",
        "work_mode":   "Remote (fully remote to start)",
        "url":         "https://himalayas.app/companies/myadvice/jobs/ai-engineer-l-generative-ai",
        "score":       90,
        "rationale":   "Near-perfect keyword match: LangChain, LlamaIndex, RAG, vector databases, prompt "
                        "engineering, AI agents/guardrails, Docker/Kubernetes, GCP all map directly onto "
                        "SAHAYAK-AI and GCP-GraphRAG; India-only, salary band (12-15 LPA) close to expected "
                        "CTC; title and scope are an exact fit for target 'Generative AI Engineer'.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 90). JD asks for 1-3 yrs experience — candidate is "
                        "significantly senior for the nominal band, positioned in materials as bringing "
                        "production-scale GenAI experience rather than overstating junior fit. LLMOps "
                        "tooling gap (MLflow/Sagemaker/W&B vs Ollama-based pipelines) flagged honestly in "
                        "cover letter. Role may relocate to Bangalore office later; candidate open to relocation.",
        "resume":      "jobs/MyAdvice_AIEngineerIGenerativeAI/MyAdvice_AIEngineerIGenerativeAI_resume.docx",
        "cover":       "jobs/MyAdvice_AIEngineerIGenerativeAI/MyAdvice_AIEngineerIGenerativeAI_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-052",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "InnovationTeam",
        "role":        "Senior Generative AI (GenAI) Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/innovationteam/jobs/senior-generative-ai-gen-ai-engineer",
        "score":       88,
        "rationale":   "Strong senior-level GenAI match: LLMs, RAG, vector databases, transformer "
                        "architectures, Docker/Kubernetes, multi-cloud deployment, and AI governance map "
                        "closely onto SAHAYAK-AI and GCP-GraphRAG; India-only remote; 5+ yr bar and "
                        "postgraduate-degree preference comfortably met by 12+ yr profile and M.Tech.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second pick this run (score 88). Vector-database experience is via LlamaIndex/Neo4j "
                        "rather than the JD's named FAISS/OpenSearch/Pinecone — flagged honestly in cover "
                        "letter. InnovationTeam is a Saudi Arabia-headquartered IT consulting firm serving "
                        "enterprise/government clients, not an in-house AI product company.",
        "resume":      "jobs/InnovationTeam_SeniorGenAIEngineer/InnovationTeam_SeniorGenAIEngineer_resume.docx",
        "cover":       "jobs/InnovationTeam_SeniorGenAIEngineer/InnovationTeam_SeniorGenAIEngineer_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-053",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Nagarro",
        "role":        "Associate Principal Engineer, Machine Learning",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/nagarro/jobs/associate-principal-engineer-machine-learning",
        "score":       75,
        "rationale":   "Strong technical overlap — NLP, GenAI, agentic multi-agent systems, RAG, GANs/VAEs, "
                        "MLOps, Docker/Kubernetes, SQL all mirror profile skills — but JD requires 9+ years "
                        "in a narrowly ML-architecture-lead capacity; above threshold but not selected as a "
                        "top-2 pick this run given the very senior experience bar relative to peer roles.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 75 — above threshold, not tailored (only top 2 tailored per run). Flag: 9+ "
                        "years total experience required in an architecture/design-authority role; profile "
                        "experience is strong but concentrated more in applied engineering than formal "
                        "architecture leadership — worth a resume pass if reconsidered next run.",
        "resume":      "",
        "cover":       "",
        "fill":        YELLOW,
    },
    {
        "job_id":      "JOB-054",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Fusemachines",
        "role":        "Machine Learning Engineer / Data Scientist",
        "location":    "India",
        "work_mode":   "Remote (Full Time)",
        "url":         "https://himalayas.app/companies/fusemachines/jobs/machine-learning-engineer-data-scientist-3554750670",
        "score":       68,
        "rationale":   "Genuine core ML/data-science role (classification, time series, clustering, "
                        "PyTorch/TensorFlow, MLOps) with good Python/SQL/Docker overlap, but the JD's "
                        "emphasis is tabular ML and applied statistics rather than GenAI/LLM/RAG — agentic "
                        "development is listed only as a nice-to-have, not a core requirement.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 68 — below 70 threshold, not tailored. Has a genuine AI/ML component so not a "
                        "deal-breaker. Would need a resume reframed toward classical ML/statistics rather "
                        "than GenAI specialisation to be competitive here.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-055",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Sabre Corporation",
        "role":        "Principal Data Science Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/sabre-corporation/jobs/principal-data-science-engineer",
        "score":       62,
        "rationale":   "Legitimate ML/statistics role (Python, SQL, TensorFlow, GCP Vertex AI, MLOps CI/CD) "
                        "with decent cloud/engineering overlap, but core focus is airline revenue-management "
                        "statistical modelling with no GenAI, LLM, RAG, or agentic-AI component in the "
                        "actual requirements — 'agentic revolution in travel' is company marketing language, "
                        "not a job requirement.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 62 — below threshold, not tailored. Requires an advanced degree in Statistics/"
                        "Operations Research/Math and domain knowledge of airline pricing/revenue management "
                        "profile does not have — significant reframing would be needed to be competitive.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-051 through JOB-055 appended to {WB_PATH}")
