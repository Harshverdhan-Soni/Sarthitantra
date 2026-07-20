"""
Daily run (2026-07-17) - append this run's 5 sourced roles to applications_tracker_NEW.xlsx. (Run 20)

  * Resolves the tracker via pathlib (BASE = this file's parent's parent), NOT a hardcoded path.
  * Dedupes on Job URL and Company+Role (safety net).
  * Assigns new Job IDs continuing from the highest JOB-### already in the sheet.

Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent          # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Bjak",
        "role": "Applied AI Engineer - Finance Super App (India)",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/bjak/jobs/applied-ai-engineer-finance-super-app-india",
        "score": 85,
        "rationale": "Top pick this run. India-based remote (candidates must be based in India). Near-ideal fit to "
                     "profile Sec 7/9: build AI-powered workflows, assistants, agents and automation; integrate LLMs "
                     "with internal data, APIs, documents and knowledge bases; RAG; evaluation, monitoring and "
                     "fallback flows; prototype->productionize on a Python/backend foundation. Maps directly to "
                     "SAHAYAK-AI (agentic, HITL), Finance SAHAYAK (multimodal RAG+OCR) and full-stack delivery. "
                     "Explicitly a builder (not research-only) role - Harsh's applied + delivery mix fits. Fintech/"
                     "insurance/support-automation is 'a strong advantage'; insurance-adjacent e-gov/healthcare "
                     "delivery is relevant.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-based remote, applied-AI/agentic must-have met, no sensitive fields to apply. "
                 "Strong English required (met). Fast process (online assessment/practical task -> tech -> CEO). "
                 "Tailored this run.",
        "resume": "jobs/Bjak_AppliedAIEngineer-FinanceSuperApp/Bjak_AppliedAIEngineer-FinanceSuperApp_resume.docx",
        "cover":  "jobs/Bjak_AppliedAIEngineer-FinanceSuperApp/Bjak_AppliedAIEngineer-FinanceSuperApp_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Cerence",
        "role": "Senior AI Scientist",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/cerence/jobs/senior-ai-scientist",
        "score": 63,
        "rationale": "Below threshold. India-eligible remote and AI/ML must-have present (conversational AI / LLM "
                     "platform, XUI). BUT the JD is a foundation-model pretraining research role: design and train "
                     "large-scale transformer/hybrid models from first principles, MoE, attention variants "
                     "(RoPE/ALiBi/GQA), optimizer/scheduler ownership and debugging training instabilities, "
                     "scaling laws, RLHF/DPO/GRPO, and distributed training (FSDP, ZeRO-3, tensor/pipeline "
                     "parallelism, bf16/fp8). This is a genuine gap vs Harsh's applied RAG/GraphRAG/GNN/agentic "
                     "profile - he applies transformers but has not pretrained foundation models at scale. Not "
                     "tailored.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (core requirement is large-scale foundation-model pretraining and distributed training "
                 "- FSDP/ZeRO-3, MoE, scaling laws, RLHF/DPO/GRPO - not in profile). Strong applied-AI candidate but "
                 "this is a pretraining research scientist role. Surfaced for awareness; not tailored.",
        "resume": "", "cover": "", "fill": PINK,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Honor (Honor Technology / Home Instead)",
        "role": "Senior Engineer, Applied AI",
        "location": "Remote (listed under India; comp in USD - likely US-remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/honorcare/jobs/senior-engineer-applied-ai",
        "score": 70,
        "rationale": "Strong content fit but eligibility uncertain. Applied AI team: LLM-powered products, AI agents, "
                     "retrieval/memory systems, evaluation frameworks, RBAC/auditability, human-in-the-loop, "
                     "workflow automation, Python + React - excellent overlap with SAHAYAK-AI, GCP-GraphRAG and "
                     "full-stack profile. HOWEVER hiring salary range is USD $160-180k with US 401k benefits and "
                     "'national average' pay, strongly implying a US-based remote role despite appearing in the "
                     "India AI listing. On-call rotation. Scored at threshold on fit; held from tailoring pending "
                     "location confirmation.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (verify) - USD comp ($160-180k), US 401k/benefits and 'national average' pay indicate "
                 "US-remote, not India-eligible; no sponsorship stated. Confirm whether India applicants are "
                 "accepted before applying. Excellent skills fit if location works. Not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Avalara (via Bitovi)",
        "role": "AI Workflow Automation Engineer",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/bitovi/jobs/ai-workflow-automation-engineer",
        "score": 74,
        "rationale": "Above threshold, not top 2. India-eligible remote (Bitovi screens; Avalara hires). Builds "
                     "enterprise AI-enabled automation with n8n as core platform plus LLMs, AI agents, prompt "
                     "management, human-in-the-loop and responsible-AI governance - n8n, agentic AI, MCP and prompt "
                     "engineering are all in profile Sec 7, and 10+ yrs is met by 12+. BUT the core requirement is "
                     "deep hands-on n8n/Boomi platform architecture and enterprise integration (REST/webhooks/OAuth/"
                     "JWT, Boomi, CI/CD for workflows) - narrower and more iPaaS/automation-platform-specialist than "
                     "Harsh's RAG/agentic research + full-stack strength. Solid fit, behind the two tailored roles.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (India remote; AI/agentic must-have met; n8n and agentic AI in profile; 10+ yrs met "
                 "via 12+). Deep n8n/Boomi platform-architect specialism is a partial gap vs research focus. Above "
                 "threshold - not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Credit Acceptance",
        "role": "Senior Software Engineer, Enterprise AI Enablement (EoR India)",
        "location": "India only (Remote)", "work_mode": "Remote (US business-hours overlap)",
        "url": "https://himalayas.app/companies/credit-acceptance/jobs/senior-software-engineer-enterprise-ai-enablement-employer-of-record",
        "score": 87,
        "rationale": "Second tailored pick (highest fit this run). Explicitly 'remote from India' via EoR partner; "
                     "CTC INR 38.8L-56.9L (well above expected 15LPA). Build applied AI across document intelligence, "
                     "decision support, fraud intelligence and intelligent assistants; implement context-aware "
                     "agentic systems orchestrating tools/APIs/reasoning; strong SWE fundamentals - OOP languages & "
                     "design patterns, system design, APIs, CI/CD, data modeling; mission-critical enterprise apps; "
                     "mentoring; LLM-powered apps, prompt/context engineering, evaluation/guardrails/compliance. Maps "
                     "cleanly to Harsh's Java/Spring Boot OOP at national scale + SAHAYAK-AI (agentic, HITL) + Finance "
                     "SAHAYAK (document intelligence RAG+OCR). 5+ yrs required vs 12+.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-remote via EoR, applied-AI/agentic must-have met, no sensitive fields to apply. "
                 "Minor friction: regular overlap with U.S. business hours + on-call rotation - confirm Harsh can "
                 "accommodate. Tailored this run.",
        "resume": "jobs/CreditAcceptance_SrSWE-EnterpriseAIEnablement/CreditAcceptance_SrSWE-EnterpriseAIEnablement_resume.docx",
        "cover":  "jobs/CreditAcceptance_SrSWE-EnterpriseAIEnablement/CreditAcceptance_SrSWE-EnterpriseAIEnablement_cover.docx",
        "fill": GREEN,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

existing_urls, existing_pairs = set(), set()
max_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    jid = str(row[0]).strip() if row[0] else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
    company = (str(row[3]).strip().lower() if len(row) > 3 and row[3] else "")
    role    = (str(row[4]).strip().lower() if len(row) > 4 and row[4] else "")
    url     = (str(row[7]).strip().lower() if len(row) > 7 and row[7] else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

next_id = max_id + 1
appended, skipped = [], []

for job in jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        skipped.append(f'{job["company"]} - {job["role"]} (already in tracker)')
        continue

    job_id = f"JOB-{next_id:03d}"
    next_id += 1

    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    appended.append(f'{job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker updated:", WB_PATH)
for a in appended:
    print("  appended", a)
for s in skipped:
    print("  skipped ", s)
