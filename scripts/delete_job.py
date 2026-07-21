"""
delete_job.py - permanently DELETE a job from the tracker list.

This removes the job's row from applications_tracker_NEW.xlsx and drops it from
apply_queue.json. Use this when you never want to see a role again. (If you just
want to stop a scheduled run from applying, use cancel_application.py instead - that
keeps the row for your records.)

Usage (from the scripts/ folder):
    python delete_job.py JOB-095                 -> asks for y/N confirmation
    python delete_job.py JOB-095 --yes           -> delete without prompting
    python delete_job.py JOB-095 JOB-094 --yes   -> delete several at once
    python delete_job.py JOB-095 --files --yes   -> also delete its jobs/<folder> docs

Deletion is irreversible, so it confirms first unless you pass --yes.
"""
import sys
import re
import json
import shutil
import pathlib
from openpyxl import load_workbook

import json as _json

BASE = pathlib.Path(__file__).parent.parent
QUEUE_PATH = BASE / "apply_queue.json"
JOBS_DIR   = BASE / "jobs"

def _tracker_path():
    at = BASE / "active_track.json"
    if at.exists():
        try:
            data = _json.loads(at.read_text(encoding="utf-8"))
            return str(BASE / data.get("tracker_file", "applications_tracker.xlsx"))
        except Exception:
            pass
    return str(BASE / "applications_tracker.xlsx")

WB_PATH = _tracker_path()

COL_ID, COL_COMPANY, COL_ROLE, COL_RESUME = 1, 4, 5, 14


def parse_args(argv):
    job_ids = []
    assume_yes = "--yes" in argv
    del_files = "--files" in argv
    for a in argv:
        m = re.match(r"(?i)job-?(\d+)", a)
        if m:
            job_ids.append("JOB-%03d" % int(m.group(1)))
    return job_ids, assume_yes, del_files


def find_rows(ws, job_ids):
    hits = {}
    for r in range(2, ws.max_row + 1):
        jid = str(ws.cell(r, COL_ID).value).strip().upper()
        if jid in set(job_ids):
            hits[jid] = (r, ws.cell(r, COL_COMPANY).value, ws.cell(r, COL_ROLE).value,
                         ws.cell(r, COL_RESUME).value)
    return hits


def main():
    job_ids, assume_yes, del_files = parse_args(sys.argv[1:])
    if not job_ids:
        try:
            entered = input("Enter the Job ID to DELETE (e.g. JOB-095): ").strip()
        except EOFError:
            entered = ""
        m = re.match(r"(?i)job-?(\d+)", entered)
        if not m:
            print("No valid Job ID given. Nothing changed.")
            return
        job_ids = ["JOB-%03d" % int(m.group(1))]

    wb = load_workbook(WB_PATH)
    ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
    hits = find_rows(ws, job_ids)

    missing = [j for j in job_ids if j not in hits]
    for j in missing:
        print("  " + j + ": NOT FOUND in tracker - skipped.")
    if not hits:
        print("Nothing to delete.")
        return

    print("About to permanently DELETE these rows from the tracker:")
    for jid, (r, company, role, _) in hits.items():
        print("  [" + jid + "] " + str(company) + " - " + str(role))
    if del_files:
        print("  ...and their jobs/<folder> documents (--files).")

    if not assume_yes:
        try:
            ans = input("Type 'yes' to confirm deletion: ").strip().lower()
        except EOFError:
            ans = ""
        if ans != "yes":
            print("Cancelled - nothing deleted.")
            return

    # delete rows bottom-up so indices stay valid
    for jid in sorted(hits, key=lambda k: hits[k][0], reverse=True):
        ws.delete_rows(hits[jid][0], 1)
    wb.save(WB_PATH)

    # prune apply_queue.json
    removed_q = 0
    if QUEUE_PATH.exists():
        try:
            data = json.load(open(QUEUE_PATH))
            if isinstance(data, list):
                before = len(data)
                data = [j for j in data if str(j.get("job_id", "")).upper() not in hits]
                json.dump(data, open(QUEUE_PATH, "w"), indent=2)
                removed_q = before - len(data)
        except Exception:
            pass

    # optionally delete the job's document folder
    removed_dirs = []
    if del_files:
        for jid, (_, _, _, resume) in hits.items():
            folder = None
            if resume:
                # resume path looks like jobs/<Folder>/<file>.docx
                parts = pathlib.Path(str(resume)).parts
                if "jobs" in parts:
                    idx = parts.index("jobs")
                    if idx + 1 < len(parts):
                        folder = JOBS_DIR / parts[idx + 1]
            if folder and folder.exists():
                shutil.rmtree(folder, ignore_errors=True)
                removed_dirs.append(str(folder))

    print("Deleted", len(hits), "row(s) from the tracker.")
    print("Removed from apply_queue.json:", removed_q, "entry(ies).")
    if del_files:
        for d in removed_dirs:
            print("  removed folder:", d)


if __name__ == "__main__":
    main()
