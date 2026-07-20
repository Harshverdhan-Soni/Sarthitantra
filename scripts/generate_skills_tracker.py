"""
Generate skills_development_tracker.html — a self-contained, offline page that,
for every active (non-skipped) job in applications_tracker_NEW.xlsx, lists the
specific skill/knowledge gaps identified during scoring, with learning
resources and a checkbox. Progress is shared across jobs (one skill, checked
once, applies everywhere it's needed) and is saved to the browser's
localStorage automatically, plus an Export/Import JSON button pair for backup
/ portability between browsers or machines.

Re-run this script after future sourcing runs to refresh the job list. Your
checked-off progress is preserved (localStorage persists across regenerations
of this file at the same path; use Export before switching browsers/machines
and Import after).

Usage:
    python generate_skills_tracker.py
"""
import json
import pathlib
import re
from datetime import date

import openpyxl

BASE = pathlib.Path(__file__).parent.parent  # JobFinder root
TRACKER = BASE / "applications_tracker_NEW.xlsx"
OUT = BASE / "skills_development_tracker.html"

# ─────────────────────────────────────────────────────────────────────────
# Skill catalog: id -> title, plain-language "why it matters", resources,
# rough time estimate. Resources are stable, well-known, mostly-free sources.
# ─────────────────────────────────────────────────────────────────────────
SKILLS = {
    "lora_qlora": {
        "title": "LLM fine-tuning (LoRA / QLoRA / PEFT)",
        "why": "Parameter-efficient fine-tuning is the standard way teams adapt open-weight "
               "LLMs to a domain without full retraining. Several JDs list this explicitly.",
        "time": "1-2 weekends (hands-on)",
        "resources": [
            {"label": "Hugging Face PEFT docs", "url": "https://huggingface.co/docs/peft/en/index"},
            {"label": "DeepLearning.AI — Finetuning Large Language Models", "url": "https://www.deeplearning.ai/short-courses/finetuning-large-language-models/"},
        ],
    },
    "post_training": {
        "title": "Post-training methods (SFT, DPO, RLHF, GRPO)",
        "why": "Increasingly asked for in applied-AI-research roles building agentic/instruction-tuned "
               "systems, e.g. Tether, PlexTrac.",
        "time": "1-2 weeks (concepts + a small hands-on run)",
        "resources": [
            {"label": "Hugging Face TRL docs (SFT/DPO trainers)", "url": "https://huggingface.co/docs/trl/en/index"},
            {"label": "DeepLearning.AI — Reinforcement Learning from Human Feedback", "url": "https://www.deeplearning.ai/short-courses/reinforcement-learning-from-human-feedback/"},
        ],
    },
    "distributed_training": {
        "title": "Distributed training (FSDP, DeepSpeed, Accelerate)",
        "why": "Research-heavy roles (Tether, Featherless AI distillation role) expect multi-GPU "
               "training experience beyond single-node fine-tuning.",
        "time": "1-2 weeks",
        "resources": [
            {"label": "PyTorch FSDP tutorial", "url": "https://pytorch.org/tutorials/intermediate/FSDP_tutorial.html"},
            {"label": "DeepSpeed — Getting Started", "url": "https://www.deepspeed.ai/getting-started/"},
            {"label": "Hugging Face Accelerate docs", "url": "https://huggingface.co/docs/accelerate/en/index"},
        ],
    },
    "dspy_prompt_opt": {
        "title": "Prompt-optimization frameworks (DSPy or similar)",
        "why": "PlexTrac and similar roles ask for systematic prompt optimization rather than "
               "manual iteration.",
        "time": "2-3 days",
        "resources": [{"label": "DSPy documentation", "url": "https://dspy.ai/"}],
    },
    "azure_ai": {
        "title": "Azure AI stack (Azure OpenAI, AI Search, Prompt Flow)",
        "why": "NorthBay (Azure AI Specialist) and Xenon7 list Azure as the core cloud platform; "
               "your production experience is on-premise/Docker-Kubernetes instead.",
        "time": "1 week (fundamentals) + hands-on lab",
        "resources": [
            {"label": "Microsoft Learn — AI on Azure learning path", "url": "https://learn.microsoft.com/en-us/training/paths/get-started-with-artificial-intelligence-on-azure/"},
            {"label": "Azure AI Search documentation", "url": "https://learn.microsoft.com/en-us/azure/search/"},
        ],
    },
    "aws_cloud": {
        "title": "AWS AI stack (SageMaker, Bedrock)",
        "why": "AHEAD, Sia Partners, and Xenon7 all expect hands-on AWS ownership (SageMaker/Bedrock) "
               "for production AI systems.",
        "time": "1 week (fundamentals) + hands-on lab",
        "resources": [
            {"label": "AWS SageMaker — Getting Started", "url": "https://aws.amazon.com/sagemaker/getting-started/"},
            {"label": "Amazon Bedrock documentation", "url": "https://docs.aws.amazon.com/bedrock/"},
        ],
    },
    "vector_db_alt": {
        "title": "Alternate vector databases (Pinecone, Weaviate, Faiss)",
        "why": "Your RAG work is Neo4j-based; SS&C and similar JDs specifically name Pinecone/Weaviate/Faiss.",
        "time": "2-3 days",
        "resources": [
            {"label": "Pinecone Learning Center", "url": "https://www.pinecone.io/learn/"},
            {"label": "Weaviate Academy", "url": "https://weaviate.io/developers/academy"},
        ],
    },
    "asr_speech": {
        "title": "Speech / ASR (Whisper, Wav2Vec, ESPnet)",
        "why": "Wadhwani AI (ASR) and Sun King both need production speech-model experience, "
               "which is not currently in your project history.",
        "time": "2-3 weeks (this is a real specialization, not a quick add-on)",
        "resources": [
            {"label": "Hugging Face Audio Course", "url": "https://huggingface.co/learn/audio-course"},
            {"label": "OpenAI Whisper (GitHub)", "url": "https://github.com/openai/whisper"},
        ],
    },
    "computer_vision": {
        "title": "Computer vision (detection/segmentation, YOLO)",
        "why": "Sun King's role pairs speech work with production CV (YOLO, detection/segmentation).",
        "time": "2-3 weeks",
        "resources": [{"label": "Hugging Face Computer Vision Course", "url": "https://huggingface.co/learn/computer-vision-course"}],
    },
    "jax_rl": {
        "title": "JAX & reinforcement learning",
        "why": "Google DeepMind's AQUA role is RL/JAX-first — a deep specialization gap from your "
               "GenAI/RAG background.",
        "time": "3-4 weeks (substantial)",
        "resources": [
            {"label": "JAX documentation", "url": "https://jax.readthedocs.io/"},
            {"label": "Hugging Face Deep RL Course", "url": "https://huggingface.co/learn/deep-rl-course"},
        ],
    },
    "gan_vae": {
        "title": "GANs & VAEs (generative modeling beyond LLMs)",
        "why": "Nagarro's Principal ML role lists GAN/VAE as a nice-to-have alongside RAG.",
        "time": "2 weeks",
        "resources": [{"label": "DeepLearning.AI — GANs Specialization", "url": "https://www.deeplearning.ai/courses/generative-adversarial-networks-gans-specialization/"}],
    },
    "mlops_platforms": {
        "title": "MLOps platforms (MLflow, Kubeflow)",
        "why": "Several ML-engineering JDs (Nagarro, Uvation) name MLflow/Kubeflow specifically; "
               "you have Docker/Kubernetes but not these ML-specific orchestration layers.",
        "time": "3-5 days",
        "resources": [
            {"label": "MLflow documentation", "url": "https://mlflow.org/docs/latest/index.html"},
            {"label": "Kubeflow documentation", "url": "https://www.kubeflow.org/docs/"},
        ],
    },
    "terraform_iac": {
        "title": "Infrastructure-as-Code (Terraform)",
        "why": "Sia Partners leans on Terraform/IaC for cloud-native microservices ownership.",
        "time": "3-5 days",
        "resources": [{"label": "Terraform — Get Started tutorials", "url": "https://developer.hashicorp.com/terraform/tutorials"}],
    },
    "databricks": {
        "title": "Databricks",
        "why": "ONE (OnePay) names Databricks as part of its data/ML stack.",
        "time": "2-3 days",
        "resources": [{"label": "Databricks Academy (free courses)", "url": "https://www.databricks.com/learn/training/home"}],
    },
    "distillation": {
        "title": "Model distillation & compression (teacher-student pipelines)",
        "why": "Featherless AI's distillation role is a specialised research area not covered by "
               "your current RAG/agentic-AI project history.",
        "time": "1-2 weeks",
        "resources": [{"label": "\"Distilling the Knowledge in a Neural Network\" (Hinton et al., foundational paper)", "url": "https://arxiv.org/abs/1503.02531"}],
    },
    "fintech_domain": {
        "title": "Fintech domain literacy (fraud, payments, claims)",
        "why": "OnePay, Luma Financial, and Cotiviti all sit in fintech/payments/claims — a domain "
               "step away from your healthcare/e-governance background.",
        "time": "1 week (self-study, domain vocabulary + case studies)",
        "resources": [{"label": "Coursera (Univ. of Michigan) — Fintech: Foundations & Applications", "url": "https://www.coursera.org/learn/fintech-foundations"}],
    },
    "csharp_dotnet": {
        "title": "C# / .NET",
        "why": "Built In's role has a C#/.NET component alongside the agentic-AI work.",
        "time": "1 week (fundamentals, if coming from Java background)",
        "resources": [{"label": "Microsoft Learn — C# first steps", "url": "https://learn.microsoft.com/en-us/training/paths/csharp-first-steps/"}],
    },
    "top_tier_pubs": {
        "title": "Top-tier publication venues (NeurIPS/ICML/ICLR/ACL/EMNLP)",
        "why": "Featherless AI and Tether expect these venues specifically; GCON 2026 (IEEE) is a "
               "respected but different tier. Strategy: submit a workshop paper or strong arXiv "
               "preprint of GCP-GraphRAG follow-up work to build toward this.",
        "time": "Ongoing — plan around your next paper cycle",
        "resources": [
            {"label": "ACL Anthology (browse target venues/CFPs)", "url": "https://aclanthology.org/"},
            {"label": "NeurIPS conference site", "url": "https://nips.cc/"},
        ],
    },
    "orchestration_frameworks": {
        "title": "Agent-orchestration frameworks (AutoGPT, Semantic Kernel)",
        "why": "NorthBay's Generative AI Lead role names these by name alongside LangChain/CrewAI, "
               "which your SAHAYAK-AI architecture already covers conceptually.",
        "time": "2-3 days (framework-specific syntax, concepts already known)",
        "resources": [
            {"label": "Microsoft Semantic Kernel docs", "url": "https://learn.microsoft.com/en-us/semantic-kernel/overview/"},
            {"label": "AutoGPT (GitHub)", "url": "https://github.com/Significant-Gravitas/AutoGPT"},
        ],
    },
    "extreme_scale_systems": {
        "title": "Extreme-scale/low-latency system design",
        "why": "HighLevel's JD operates at billions of API hits/day across 250+ microservices — "
               "a different scale tier from national-government-scale delivery.",
        "time": "1-2 weeks reading + case studies",
        "resources": [{"label": "System Design Primer (GitHub)", "url": "https://github.com/donnemartin/system-design-primer"}],
    },
    "classical_ml": {
        "title": "Classical ML (regression, clustering, time-series)",
        "why": "Uvation and Fusemachines lean on classical ML/statistical modeling more than GenAI.",
        "time": "1-2 weeks (refresher)",
        "resources": [{"label": "Coursera — Machine Learning Specialization (Andrew Ng)", "url": "https://www.coursera.org/specializations/machine-learning-introduction"}],
    },
    "security_data": {
        "title": "Security telemetry (SIEM, network/endpoint logs)",
        "why": "PlexTrac's product is cybersecurity-specific; you have zero direct experience with "
               "SIEM/network-telemetry data per profile.md.",
        "time": "3-5 days (conceptual familiarity, enough to speak to it in an interview)",
        "resources": [{"label": "MITRE ATT&CK framework overview (widely used SIEM/detection reference)", "url": "https://attack.mitre.org/"}],
    },
    "clinical_imaging": {
        "title": "Clinical imaging research (PET/CT, MRI, nuclear medicine)",
        "why": "Nucs AI requires a PhD specifically in medical-imaging research with publications in "
               "clinical-imaging venues. This is a multi-year specialization, not a course you can "
               "add quickly — flagging honestly rather than suggesting a shortcut. Only pursue this "
               "if you're genuinely interested in pivoting your PhD focus; otherwise treat JOB-043 "
               "as a long shot regardless of skill-building.",
        "time": "Not realistically closeable short-term",
        "resources": [{"label": "(No shortcut resource — this needs a genuine research-focus pivot)", "url": "https://www.nucs.ai/"}],
    },
    "review_jd": {
        "title": "Re-read the full live JD & tune resume emphasis",
        "why": "No specific hard skill gap was flagged during scoring for this role beyond fit/seniority "
               "considerations — the main lever here is making sure the tailored resume's emphasis "
               "still matches the live posting before you apply.",
        "time": "15-20 minutes",
        "resources": [],
    },
}

# ─────────────────────────────────────────────────────────────────────────
# Keyword → skill_id matching rules, run against "rationale + notes" text.
# Order matters only in that all matches are kept (deduplicated) per job.
# ─────────────────────────────────────────────────────────────────────────
RULES = [
    (r"\bC#/\.NET\b|\bC#\b", "csharp_dotnet"),
    (r"ACL/EMNLP|NeurIPS/ICML|NeurIPS|ICML|ICLR|top-tier ML conference|top-tier NLP", "top_tier_pubs"),
    (r"crypto|fintech domain|fraud, payments|payment/claims|claims-specific fintech", "fintech_domain"),
    (r"[Dd]istributed (GPU )?training at scale|[Dd]istributed[- ]training", "distributed_training"),
    (r"\bRL\b|\bJAX\b", "jax_rl"),
    (r"AutoGPT|Semantic Kernel", "orchestration_frameworks"),
    (r"Azure[- ]specific|Azure AI Search|Prompt Flow|Azure OpenAI|AWS[/,]\s?(GCP[/,]\s?)?Azure|AWS/Azure ownership", "azure_ai"),
    (r"[Dd]istillation|teacher-student|model compression", "distillation"),
    (r"ASR|speech recognition|Wav2Vec|Whisper", "asr_speech"),
    (r"computer vision|YOLO|CV\b", "computer_vision"),
    (r"Databricks", "databricks"),
    (r"Pinecone|Weaviate|Faiss", "vector_db_alt"),
    (r"AWS SageMaker|NVIDIA DGX|SageMaker|Bedrock|AWS[/,]\s?(GCP[/,]\s?)?Azure|AWS/Azure ownership", "aws_cloud"),
    (r"GAN/VAE|GAN\b|VAE\b", "gan_vae"),
    (r"MLflow|Kubeflow", "mlops_platforms"),
    (r"EdTech|K-12", "review_jd"),
    (r"Terraform|IaC\b", "terraform_iac"),
    (r"billions of API hits|250\+ microservices|extreme scale|low-latency engineering", "extreme_scale_systems"),
    (r"SIEM|network-telemetry|network telemetry|security data", "security_data"),
    (r"medical imaging|clinical-imaging|PET/CT|MRI|PSMA-PET|SPECT|nuclear-medicine", "clinical_imaging"),
    (r"classical.{0,20}ML|[Tt]raditional ML|regression/clustering|tabular/time-series|classical statistics", "classical_ml"),
    (r"LoRA|QLoRA|fine-tuning LLMs", "lora_qlora"),
    (r"SFT, DPO, RLHF|DPO/RLHF|post-training", "post_training"),
    (r"DSPy", "dspy_prompt_opt"),
]

SKIP_STATUSES = {"Skipped"}


def load_jobs():
    wb = openpyxl.load_workbook(TRACKER)
    ws = wb["Applications"]
    headers = [c.value for c in ws[1]]
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        if d.get("Status") in SKIP_STATUSES:
            continue
        text = " ".join(str(d.get(k) or "") for k in ("Score rationale", "Missing requirements / notes"))
        matched = []
        for pattern, skill_id in RULES:
            if re.search(pattern, text, re.IGNORECASE) and skill_id not in matched:
                matched.append(skill_id)
        if not matched:
            matched = ["review_jd"]
        resume_val = d.get("Resume file")
        cover_val = d.get("Cover letter file")
        jobs.append({
            "id": d.get("Job ID"),
            "company": d.get("Company"),
            "role": d.get("Role"),
            "score": d.get("Fit score"),
            "status": d.get("Status"),
            "url": d.get("Job URL"),
            "resume": resume_val if isinstance(resume_val, str) and resume_val.endswith(".docx") else None,
            "cover": cover_val if isinstance(cover_val, str) and cover_val.endswith(".docx") else None,
            "skills": matched,
        })
    jobs.sort(key=lambda j: (j["score"] is None, -(j["score"] or 0)))
    return jobs


def build_html(jobs):
    used_skill_ids = sorted({sid for j in jobs for sid in j["skills"]})
    skills_subset = {sid: SKILLS[sid] for sid in used_skill_ids}
    data = {
        "generated": date.today().isoformat(),
        "jobs": jobs,
        "skills": skills_subset,
    }
    data_json = json.dumps(data)

    html = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>JobFinder — Skills Development Tracker</title>
<style>
  :root {
    --blue: #1f4e79; --blue-light: #eaf1f8; --green: #16a766; --green-bg: #e7f7ee;
    --yellow: #b8860b; --yellow-bg: #fff8e6; --gray: #595959; --border: #e2e2e2;
    --bg: #f7f8fa;
  }
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif;
    background: var(--bg); color: #1a1a1a; margin: 0; padding: 0 0 60px 0;
  }
  header {
    background: var(--blue); color: white; padding: 24px 32px;
  }
  header h1 { margin: 0 0 4px 0; font-size: 22px; }
  header p { margin: 0; opacity: 0.9; font-size: 13.5px; }
  .wrap { max-width: 980px; margin: 0 auto; padding: 24px 20px; }
  .panel {
    background: white; border: 1px solid var(--border); border-radius: 10px;
    padding: 18px 20px; margin-bottom: 20px;
  }
  .summary-row { display: flex; gap: 16px; flex-wrap: wrap; align-items: center; }
  .stat { flex: 1; min-width: 140px; text-align: center; padding: 10px; }
  .stat .num { font-size: 26px; font-weight: 700; color: var(--blue); }
  .stat .label { font-size: 12px; color: var(--gray); }
  .toolbar { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 14px; }
  button.btn {
    background: var(--blue); color: white; border: none; border-radius: 6px;
    padding: 8px 14px; font-size: 13px; cursor: pointer;
  }
  button.btn.secondary { background: white; color: var(--blue); border: 1px solid var(--blue); }
  button.btn:hover { opacity: 0.88; }
  input[type=file] { display: none; }
  .instructions { font-size: 13px; color: var(--gray); line-height: 1.5; margin-top: 10px; }
  .job-card {
    background: white; border: 1px solid var(--border); border-radius: 10px;
    padding: 16px 20px; margin-bottom: 14px;
  }
  .job-card.ready { border-color: var(--green); box-shadow: 0 0 0 1px var(--green) inset; }
  .job-head { display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 8px; }
  .job-title { font-size: 15.5px; font-weight: 700; }
  .job-company { color: var(--gray); font-size: 13px; }
  .badge {
    font-size: 11px; font-weight: 700; padding: 3px 9px; border-radius: 999px;
    white-space: nowrap; height: fit-content;
  }
  .badge.ready { background: var(--green-bg); color: var(--green); }
  .badge.progress { background: var(--yellow-bg); color: var(--yellow); }
  .badge.score { background: var(--blue-light); color: var(--blue); margin-left: 6px; }
  .job-links { font-size: 12.5px; margin: 6px 0 10px 0; }
  .job-links a { color: var(--blue); text-decoration: none; margin-right: 14px; }
  .job-links a:hover { text-decoration: underline; }
  .progress-bar-outer {
    background: #eee; border-radius: 5px; height: 6px; margin: 8px 0 12px 0; overflow: hidden;
  }
  .progress-bar-inner { background: var(--green); height: 100%; transition: width 0.2s; }
  .skill-row {
    display: flex; align-items: flex-start; gap: 10px; padding: 8px 0;
    border-top: 1px solid #f0f0f0;
  }
  .skill-row:first-of-type { border-top: none; }
  .skill-row input[type=checkbox] { margin-top: 3px; width: 16px; height: 16px; flex-shrink: 0; }
  .skill-body { flex: 1; }
  .skill-title { font-size: 13.5px; font-weight: 600; }
  .skill-title.done { text-decoration: line-through; color: var(--gray); }
  .skill-why { font-size: 12.5px; color: var(--gray); margin: 2px 0 4px 0; line-height: 1.4; }
  .skill-meta { font-size: 11.5px; color: #888; margin-bottom: 4px; }
  .skill-resources a {
    font-size: 12px; color: var(--blue); text-decoration: none; margin-right: 12px;
  }
  .skill-resources a:hover { text-decoration: underline; }
  .footer-note { font-size: 12px; color: var(--gray); text-align: center; margin-top: 30px; }
  .status-tag { font-size: 11px; padding: 2px 8px; border-radius: 999px; background: #f0f0f0; color: #555; margin-left: 6px; }
</style>
</head>
<body>

<header>
  <h1>JobFinder — Skills Development Tracker</h1>
  <p>Generated __GENERATED__ from applications_tracker_NEW.xlsx. Check off a skill once you've genuinely
     developed/refreshed it — it applies everywhere that skill is needed. When a job turns green
     ("Ready to apply"), tell Claude the Job ID and it will proceed with the pre-fill → your final
     approval → submit workflow.</p>
</header>

<div class="wrap">

  <div class="panel">
    <div class="summary-row">
      <div class="stat"><div class="num" id="stat-skills-done">0/0</div><div class="label">Skills checked off</div></div>
      <div class="stat"><div class="num" id="stat-jobs-ready">0</div><div class="label">Jobs ready to apply</div></div>
      <div class="stat"><div class="num" id="stat-jobs-total">0</div><div class="label">Active jobs tracked</div></div>
    </div>
    <div class="toolbar">
      <button class="btn" id="export-btn">Export progress (.json)</button>
      <button class="btn secondary" id="import-btn">Import progress</button>
      <input type="file" id="import-file" accept="application/json">
      <button class="btn secondary" id="expand-all-btn">Expand all</button>
      <button class="btn secondary" id="collapse-all-btn">Collapse all</button>
    </div>
    <div class="instructions">
      Progress auto-saves to this browser (localStorage) every time you check a box. Use
      <b>Export</b> before switching browsers or computers, and <b>Import</b> to restore it. This page
      is read-only otherwise — it doesn't write back to the tracker or submit anything by itself.
      <br><br>
      <b>Linked to the master tracker:</b> open
      <a href="applications_tracker_NEW.xlsx" target="_blank">applications_tracker_NEW.xlsx</a> to see
      every job's full record (each row deep-links back here via its Job ID). To sync your
      readiness progress into that tracker's "Skill-Ready?" column: click <b>Export progress</b>
      above, save the file as <code>skills_progress_export.json</code> directly in the JobFinder
      folder (overwriting any older copy), then ask Claude to run
      <code>scripts/sync_skill_readiness.py</code>.
    </div>
  </div>

  <div id="job-list"></div>

  <div class="footer-note">
    Re-run <code>scripts/generate_skills_tracker.py</code> after future sourcing runs to refresh this
    list with new jobs — your exported progress file can be re-imported afterward.
  </div>

</div>

<script>
const DATA = __DATA_JSON__;
const STORAGE_KEY = "jobfinder_skills_progress_v1";

function loadState() {
  try {
    const raw = localStorage.getItem(STORAGE_KEY);
    return raw ? JSON.parse(raw) : {};
  } catch (e) { return {}; }
}
function saveState(state) {
  try { localStorage.setItem(STORAGE_KEY, JSON.stringify(state)); } catch (e) {}
}

let state = loadState();

function isChecked(skillId) { return !!state[skillId]; }

function jobProgress(job) {
  const total = job.skills.length;
  const done = job.skills.filter(isChecked).length;
  return { total, done, ready: total > 0 && done === total };
}

function render() {
  const list = document.getElementById("job-list");
  list.innerHTML = "";

  let readyCount = 0;
  const allSkillIds = Object.keys(DATA.skills);
  const doneSkillCount = allSkillIds.filter(isChecked).length;

  DATA.jobs.forEach(job => {
    const prog = jobProgress(job);
    if (prog.ready) readyCount++;

    const card = document.createElement("div");
    card.id = job.id;
    card.className = "job-card" + (prog.ready ? " ready" : "");

    const pct = prog.total ? Math.round((prog.done / prog.total) * 100) : 100;

    let linksHtml = `<a href="${job.url}" target="_blank" rel="noopener">Job posting ↗</a>`;
    if (job.resume) linksHtml += `<a href="${job.resume}" target="_blank">Tailored resume ↗</a>`;
    if (job.cover) linksHtml += `<a href="${job.cover}" target="_blank">Cover letter ↗</a>`;

    const skillsHtml = job.skills.map(sid => {
      const s = DATA.skills[sid];
      const checked = isChecked(sid) ? "checked" : "";
      const doneClass = isChecked(sid) ? "done" : "";
      const resourcesHtml = (s.resources || []).map(r =>
        `<a href="${r.url}" target="_blank" rel="noopener">${r.label} ↗</a>`
      ).join("");
      return `
        <div class="skill-row">
          <input type="checkbox" data-skill="${sid}" ${checked}>
          <div class="skill-body">
            <div class="skill-title ${doneClass}">${s.title}</div>
            <div class="skill-why">${s.why}</div>
            <div class="skill-meta">Est. time: ${s.time}</div>
            <div class="skill-resources">${resourcesHtml}</div>
          </div>
        </div>`;
    }).join("");

    card.innerHTML = `
      <div class="job-head">
        <div>
          <div class="job-title">${job.id} — ${job.role}</div>
          <div class="job-company">${job.company}
            <span class="status-tag">${job.status}</span>
          </div>
        </div>
        <div>
          ${job.score != null ? `<span class="badge score">Fit score ${job.score}</span>` : ''}
          ${prog.ready
            ? '<span class="badge ready">✓ Ready to apply</span>'
            : `<span class="badge progress">${prog.done}/${prog.total} skills</span>`}
        </div>
      </div>
      <div class="job-links">${linksHtml}</div>
      <div class="progress-bar-outer"><div class="progress-bar-inner" style="width:${pct}%"></div></div>
      <div class="skills-container">${skillsHtml}</div>
    `;
    list.appendChild(card);
  });

  document.getElementById("stat-skills-done").textContent = `${doneSkillCount}/${allSkillIds.length}`;
  document.getElementById("stat-jobs-ready").textContent = readyCount;
  document.getElementById("stat-jobs-total").textContent = DATA.jobs.length;

  document.querySelectorAll('input[type=checkbox][data-skill]').forEach(cb => {
    cb.addEventListener("change", (e) => {
      const sid = e.target.getAttribute("data-skill");
      state[sid] = e.target.checked;
      saveState(state);
      render();
    });
  });
}

document.getElementById("export-btn").addEventListener("click", () => {
  const blob = new Blob([JSON.stringify(state, null, 2)], { type: "application/json" });
  const url = URL.createObjectURL(blob);
  const a = document.createElement("a");
  a.href = url;
  a.download = "skills_progress_export.json";
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  URL.revokeObjectURL(url);
});

document.getElementById("import-btn").addEventListener("click", () => {
  document.getElementById("import-file").click();
});
document.getElementById("import-file").addEventListener("change", (e) => {
  const file = e.target.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = (evt) => {
    try {
      const imported = JSON.parse(evt.target.result);
      state = Object.assign({}, state, imported);
      saveState(state);
      render();
    } catch (err) {
      alert("Could not read that file as valid progress JSON.");
    }
  };
  reader.readAsText(file);
});

document.getElementById("expand-all-btn").addEventListener("click", () => {
  document.querySelectorAll(".skills-container").forEach(el => el.style.display = "block");
});
document.getElementById("collapse-all-btn").addEventListener("click", () => {
  document.querySelectorAll(".skills-container").forEach(el => el.style.display = "block");
});

render();
</script>

</body>
</html>
"""
    html = html.replace("__GENERATED__", data["generated"])
    html = html.replace("__DATA_JSON__", data_json)
    return html


def main():
    jobs = load_jobs()
    html = build_html(jobs)
    OUT.write_text(html, encoding="utf-8")
    print(f"Written: {OUT}  ({len(jobs)} active jobs)")


if __name__ == "__main__":
    main()
