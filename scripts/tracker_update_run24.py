"""
Daily run (2026-07-20) - Run 24.

Appends the 5 NEW roles sourced this run to applications_tracker_NEW.xlsx.
Tracker path resolved via pathlib (BASE = this file's parent's parent) - never hardcoded.
Dedupe is by Job URL and Job ID, so re-running is safe (already-logged rows are skipped).

Colour coding:
  GREEN  #D9EAD3 = Tailored
  YELLOW #FFF2CC = Scored / above threshold but not tailored
  PINK   #FFDACC = Below threshold or Skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import date

BASE = pathlib.Path(__file__).parent.parent
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK = PatternFill("solid", fgColor="FFDACC")
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()

new_jobs = [
    {
        "id": "JOB-116", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Globalization Partners (G-P)",
        "role": "Principal Software Engineer - AI Platform",
        "location": "India only (Remote)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/globalization-partners/jobs/principal-software-engineer-ai-platform",
        "score": 87,
        "rationale": "Top pick this run. LIVE (posted Jul 11 2026, apply before Sep 09 2026), India-only remote so no "
                     "sponsorship needed. The JD is an unusually complete match: architect customer-facing AI agents "
                     "and multi-agent 'Super Agents' that reason, plan and execute multi-step tasks; design "
                     "workflow-automation frameworks; set best practices for LLM integration, prompt engineering and "
                     "agent orchestration; and do it across a modern backend stack (Java, Node.js, Python, Golang) "
                     "with SQL and NoSQL, AI/ML frameworks (PyTorch/TensorFlow) and cloud. This spans exactly Harsh's "
                     "two strengths at once - production agentic AI (SAHAYAK-AI, GCP-GraphRAG) AND scalable "
                     "enterprise platform engineering (Java/Spring Boot national-scale delivery) - and asks 12+ yrs "
                     "with mentorship, which he has (12+ yrs; applied-AI lead; PG-DAI/PGDAC faculty). Java as a "
                     "first-class backend here is a genuine differentiator few GenAI candidates bring. Gap: "
                     "single-cloud/serverless depth vs his on-prem/containerised delivery - transferable, flagged "
                     "honestly.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, agentic-AI/LLM must-have met, no sensitive fields required to "
                 "apply. Tailored this run: resume leads with multi-agent 'super agents' + LLMs-in-production + "
                 "Java/Python/Go scalable-platform architecture + technical leadership; cover letter hooks on the "
                 "platform-engineering-plus-agentic-AI intersection and is candid on single-cloud depth. Note: "
                 "aggregator (Himalayas) link stored; confirm G-P's official careers posting as the apply source "
                 "before submitting.",
        "resume": "jobs/GlobalizationPartners_PrincipalSWE-AIPlatform/"
                  "GlobalizationPartners_PrincipalSWE-AIPlatform_resume.docx",
        "cover": "jobs/GlobalizationPartners_PrincipalSWE-AIPlatform/"
                 "GlobalizationPartners_PrincipalSWE-AIPlatform_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-117", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Unisys",
        "role": "Lead Engineer AI/ML",
        "location": "India only (Remote)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/unisys/jobs/lead-engineer-ai-ml",
        "score": 86,
        "rationale": "Second tailored pick. LIVE (posted Jul 13 2026, apply before Sep 11 2026), India-only remote. "
                     "The JD is an unusually clean match to profile Sec 7: lead the design and development of "
                     "Generative AI applications with LLMs, embeddings, vector databases, RAG and multi-agent "
                     "frameworks; solid Agentic AI (planning, reasoning, multi-agent collaboration) with "
                     "orchestration frameworks such as CrewAI; Python plus Golang API-driven back-end; PostgreSQL and "
                     "vector databases; a front-end framework such as React; cloud-native containers/microservices on "
                     "Kubernetes, Docker, Helm; and Figma for prototyping - Harsh has every one of these. His "
                     "Java/Spring Boot + ReactJS + Go + Docker/K8s full-stack delivery plus SAHAYAK-AI agentic work "
                     "is a genuine differentiator. Requirement 8+ yrs (he has 12+). Gap: the role centres on Azure "
                     "OpenAI / Azure AI Services, whereas his LLM delivery is GPT/Claude/Gemini + on-prem Ollama; "
                     "Azure cert is only 'an added advantage' - flagged honestly in the cover letter.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, GenAI/agentic must-have met, no sensitive fields required to "
                 "apply. Tailored this run: resume leads with GenAI (LLMs/RAG/vector DBs) + agentic AI + Python/Go "
                 "API-driven + React + Kubernetes/Docker + Figma; cover letter hooks on the full-stack-of-AI scope "
                 "and is candid about Azure-specific depth vs his GPT/Claude/Gemini + Ollama background. Note: "
                 "aggregator (Himalayas) link stored; confirm Unisys's official careers posting before submitting.",
        "resume": "jobs/Unisys_LeadEngineer-AIML/Unisys_LeadEngineer-AIML_resume.docx",
        "cover": "jobs/Unisys_LeadEngineer-AIML/Unisys_LeadEngineer-AIML_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-118", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Ollion",
        "role": "Lead - AI/ML Engineer (Generative AI Agent Systems)",
        "location": "India only (Remote)",
        "work_mode": "Remote (Contractor)",
        "url": "https://himalayas.app/companies/ollion/jobs/lead-ai-ml-engineer-generative-ai-agent-systems",
        "score": 74,
        "rationale": "Above threshold, NOT tailored (not top 2). LIVE (posted Jul 18 2026, apply before Sep 16 2026), "
                     "India-only remote. Strong content overlap with profile Sec 7: production multi-agent "
                     "frameworks, LLM reasoning loops, function calling, prompt engineering, human-in-the-loop, "
                     "shadow-mode evaluation and AI governance - all core SAHAYAK-AI material - plus a regulated / "
                     "public-sector data-compliance emphasis that matches his governed-government-systems background. "
                     "Held below the top 2 for two reasons: (1) it is a client-facing Principal-Consultant / SME + "
                     "pre-sales delivery-leadership role (technical discoveries, playbooks, CTO/VP stakeholders) "
                     "rather than hands-on build, which is where Harsh is strongest; and (2) it demands expert-level "
                     "Google Vertex AI Agent Development Kit (ADK) and hardened GCP security plus advanced Go as the "
                     "primary back-end - his agentic work is on-prem/Ollama and his Go is solid but not his deepest "
                     "stack. Contractor engagement.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Provisional - India-only remote and GenAI/agentic must-have clearly met, but the role "
                 "wants deep Vertex AI ADK + hardened GCP security and Go as primary back-end, plus 6-12 yrs of "
                 "professional-services / pre-sales consulting leadership - a different centre of gravity from "
                 "Harsh's applied-build profile. AI-content fit is high; role-type (consulting SME) and "
                 "Vertex/ADK/GCP specificity are the gaps. Surfaced for Harsh's decision rather than auto-tailored - "
                 "say the word and it can be tailored/pre-filled.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "id": "JOB-119", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Newfold Digital",
        "role": "Software Engineer AI",
        "location": "India only (Remote)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/newfold-digital/jobs/software-engineer-ai",
        "score": 72,
        "rationale": "Above threshold, NOT tailored (not top 2). LIVE (posted Jul 06 2026, apply before Sep 04 2026), "
                     "India-only remote. The AI content is a strong profile-Sec-7 match and includes a standout hook: "
                     "build agents and orchestrators on an agentic-AI platform, and - as a core, required "
                     "responsibility - develop and integrate MCP servers and tools (Python FastMCP). MCP Server is "
                     "explicitly in Harsh's toolkit and central to SAHAYAK-AI's MCP governance, so this is unusually "
                     "on-point; add multi-agent workflows, prompt engineering, LLM eval pipelines, RAG with vector "
                     "stores, Langfuse observability, CI/CD and Docker - all his material. Held below the top 2 "
                     "because the role is banded at 2-4 yrs of Python (mid-level despite the 'Senior' tag) and framed "
                     "as building under guidance from senior engineers - below a 12+ yr PhD-track engineer's level. "
                     "No AI/ML deal-breaker; genuinely relevant but junior-of-fit on seniority.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes on skills (MCP/FastMCP - a direct profile match - plus agents, RAG, prompt "
                 "engineering, LLM evals, Python all present) but MID-LEVEL banding (2-4 yrs, work 'with guidance "
                 "from senior engineers') sits below his seniority. Not auto-tailored - surfaced for Harsh's "
                 "decision. Strong MCP-centred hook if he wants a low-friction India-remote agentic role; can be "
                 "tailored/pre-filled on request.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "id": "JOB-120", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Tolken",
        "role": "Applied Scientist / Applied ML Engineer",
        "location": "IN and US only (Remote)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/tolken/jobs/applied-scientist-applied-ml-engineer-5804014310",
        "score": 64,
        "rationale": "BELOW THRESHOLD (70) - not tailored, not pre-filled. LIVE but CLOSING (posted May 24 2026, "
                     "apply before Jul 23 2026), India eligible (IN and US only). AI/ML must-have is met (real ML in "
                     "production, Python/pandas/NumPy/scikit-learn, PyTorch/TensorFlow, SQL, REST/gRPC, Docker, SHAP "
                     "explainability) and the shared engineering layer is solid. Scored below threshold because the "
                     "role's core is classical / optimisation ML for pricing, bidding and risk decisioning on a "
                     "cross-border payments platform - dynamic pricing, bandits, auctions/marketplace, fraud/AML/"
                     "credit-risk scoring - which is a distinct track from Harsh's GenAI / RAG / GraphRAG / "
                     "multilingual-NLP core. It is also explicitly Mid-level (3-7 yrs). Honest fit rather than a "
                     "stretch to clear the bar.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (core is dynamic-pricing / bidding / auctions / fraud-AML risk-scoring optimisation "
                 "ML, which Harsh does not have; his production ML is GenAI/NLP/GNN). AI/ML must-have is met and it "
                 "is remote incl. India, so surfaced for Harsh's decision rather than skipped outright - but it is a "
                 "genuine domain mismatch, not a threshold-shaving case, and the window closes Jul 23 2026. Revisit "
                 "only if he wants to pivot toward pricing/marketplace/risk ML.",
        "resume": "", "cover": "", "fill": PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications']

existing_urls = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}
existing_ids = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}

added = 0
for job in new_jobs:
    if job['url'] in existing_urls or job['id'] in existing_ids:
        print(f"SKIP (already tracked): {job['id']} {job['company']} - {job['role']}")
        continue
    r = ws.max_row + 1
    row = [
        job['id'], job['date'], job['source'], job['company'], job['role'],
        job['location'], job['work_mode'], job['url'], job['score'], job['rationale'],
        job['status'], job['dealbreaker'], job['notes'], job['resume'], job['cover'],
        '', '', '', TODAY, '', '', '',
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = job['fill']
        cell.alignment = WRAP
    added += 1
    print(f"ADDED {job['id']}: {job['company']} - {job['role']} (score {job['score']}, {job['status']})")

wb.save(WB_PATH)
print(f"\nSaved {WB_PATH} - {added} row(s) appended. Last row now {ws.max_row}.")
