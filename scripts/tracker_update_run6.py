"""
Append JOB-026 through JOB-030 (Run 6 — 2026-07-03) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-026",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Nagarro",
        "role":        "Principal Engineer, Machine Learning",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/nagarro/jobs/principal-engineer-machine-learning",
        "score":       87,
        "rationale":   "Near-perfect match: multi-agent orchestration for autonomous decision-making, LLM/foundation "
                        "model prompt engineering, and RAG map directly onto SAHAYAK-AI and GCP-GraphRAG; 11+ yrs "
                        "bar fits 12+ yrs profile exactly; India remote, 5000+ employee stable employer.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 87). Gap: no direct GAN/VAE or JAX experience in profile (JD lists "
                        "these as nice-to-have alongside RAG); MLflow/Kubeflow not explicit vs. Docker/Kubernetes "
                        "which profile does have.",
        "resume":      "jobs/Nagarro_PrincipalEngineerMachineLearning/Nagarro_PrincipalEngineerMachineLearning_resume.docx",
        "cover":       "jobs/Nagarro_PrincipalEngineerMachineLearning/Nagarro_PrincipalEngineerMachineLearning_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-027",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Nagarro",
        "role":        "Associate Principal Engineer, Machine Learning",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/nagarro/jobs/associate-principal-engineer-machine-learning",
        "score":       84,
        "rationale":   "Same JD/skill stack as JOB-026 (multi-agent systems, LLM/RAG, MLOps) at a lower seniority "
                        "tier (9+ yrs vs 11+ yrs) — strong fit but redundant with the Principal-level role at the "
                        "same company already tailored this run.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 84 — above threshold but not tailored (duplicate opportunity at same company as "
                        "JOB-026, only one Nagarro application queued this run per daily-cap discipline).",
        "resume":      "",
        "cover":       "",
        "fill":        YELLOW,
    },
    {
        "job_id":      "JOB-028",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "AHEAD",
        "role":        "Senior Technical Consultant - Data and AI",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/ahead/jobs/senior-technical-consultant-data-and-ai",
        "score":       82,
        "rationale":   "Strong GenAI/LLM/agentic-model and RAG fit combined with client-facing consulting, pre-sales "
                        "scoping, and mentoring — mirrors PMU coordination and PG-DAI/PGDAC faculty experience; "
                        "5+ yrs bar comfortably met by 12+ yrs profile.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second top pick (score 82). Gap: hands-on experience is on-premise (Ollama) rather than "
                        "Azure OpenAI / AWS SageMaker / NVIDIA DGX specifically — cover letter flags this honestly "
                        "while noting transferable RAG/fine-tuning/prompt-engineering concepts.",
        "resume":      "jobs/AHEAD_SeniorTechnicalConsultant-DataAI/AHEAD_SeniorTechnicalConsultant-DataAI_resume.docx",
        "cover":       "jobs/AHEAD_SeniorTechnicalConsultant-DataAI/AHEAD_SeniorTechnicalConsultant-DataAI_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-029",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Fusemachines",
        "role":        "Machine Learning Engineer / Data Scientist",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/fusemachines/jobs/machine-learning-engineer-data-scientist-3554750670",
        "score":       66,
        "rationale":   "Core tabular/time-series/clustering ML and deep-learning role for an AI services firm; "
                        "generative AI and agentic development are listed only as nice-to-have, not the central "
                        "ask — weaker fit than profile's GenAI/RAG specialisation.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 66 — below 70 threshold. Not tailored. Has a genuine AI/ML component so not a "
                        "deal-breaker, just a classical-ML fit rather than profile's GenAI/agentic strength.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-030",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Sabre Corporation",
        "role":        "Principal Data Science Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/sabre-corporation/jobs/principal-data-science-engineer",
        "score":       60,
        "rationale":   "Classical statistics/operations-research-heavy data science for airline pricing and revenue "
                        "management; no GenAI/LLM component in requirements; domain (airline/travel) and cloud "
                        "tooling (GCP Vertex AI, Terraform) diverge from profile's healthcare/e-gov GenAI focus.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 60 — below 70 threshold. Not tailored. Has an AI/ML component (ML/statistical "
                        "modeling) so not a deal-breaker, just a weak domain and GenAI-specialisation fit.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-026 through JOB-030 appended to {WB_PATH}")
