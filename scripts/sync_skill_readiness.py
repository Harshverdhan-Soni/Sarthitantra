"""
Sync 'Skill-Ready?' status from the exported skills-tracker progress JSON
back into applications_tracker_NEW.xlsx, and add a deep-link column back to
skills_development_tracker.html for each job.

Workflow:
  1. Open skills_development_tracker.html, check off skills as you develop them.
  2. Click "Export progress (.json)" -> save the downloaded file as
     skills_progress_export.json directly in the JobFinder root folder
     (overwrite if it already exists).
  3. Run this script (or ask Claude to run it):  python sync_skill_readiness.py

If no export file exists yet, this script still runs and treats every skill
as not-yet-developed (so "Skill-Ready?" = No and "Skill gaps remaining" lists
everything) -- safe to run any time, including right after sourcing new jobs.

Adds three columns to the Applications sheet if they don't already exist:
    - "Skill-Ready?"          Yes / No
    - "Skill gaps remaining"  semicolon-separated list of skill titles still unchecked
    - "Skills page"           hyperlink to skills_development_tracker.html#<Job ID>

Usage:
    python sync_skill_readiness.py [path-to-progress-json]
"""
import json
import pathlib
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.hyperlink import Hyperlink

from generate_skills_tracker import SKILLS, load_jobs, TRACKER, BASE

DEFAULT_PROGRESS = BASE / "skills_progress_export.json"
SKILLS_HTML = "skills_development_tracker.html"

READY_FILL = PatternFill("solid", fgColor="D9EAD3")
NOT_READY_FILL = PatternFill("solid", fgColor="FFF2CC")
WRAP = Alignment(wrap_text=True, vertical="top")
LINK_FONT = Font(color="1155CC", underline="single")

COL_READY = "Skill-Ready?"
COL_GAPS = "Skill gaps remaining"
COL_LINK = "Skills page"


def main():
    progress_path = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PROGRESS
    if progress_path.exists():
        state = json.loads(progress_path.read_text(encoding="utf-8"))
        print(f"Loaded progress from {progress_path.name} ({sum(1 for v in state.values() if v)} skills checked).")
    else:
        state = {}
        print(f"No progress file found at {progress_path.name} -- treating all skills as not yet developed. "
              f"Export from skills_development_tracker.html and re-run to get real status.")

    jobs = load_jobs()  # active (non-skipped) jobs w/ matched skill ids, from live tracker

    wb = openpyxl.load_workbook(TRACKER)
    ws = wb["Applications"]
    headers = [c.value for c in ws[1]]

    for col_name in (COL_READY, COL_GAPS, COL_LINK):
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name).font = Font(bold=True)

    col_ready_idx = headers.index(COL_READY) + 1
    col_gaps_idx = headers.index(COL_GAPS) + 1
    col_link_idx = headers.index(COL_LINK) + 1
    col_jobid_idx = headers.index("Job ID") + 1

    # widen the new columns a bit so they're readable
    for idx, width in ((col_ready_idx, 13), (col_gaps_idx, 45), (col_link_idx, 16)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    jobs_by_id = {j["id"]: j for j in jobs}
    updated = 0
    ready_count = 0

    for row in range(2, ws.max_row + 1):
        job_id = ws.cell(row=row, column=col_jobid_idx).value
        if job_id not in jobs_by_id:
            continue  # skipped / deal-breaker rows aren't tracked here
        job = jobs_by_id[job_id]
        remaining = [SKILLS[sid]["title"] for sid in job["skills"] if not state.get(sid)]
        ready = len(remaining) == 0
        if ready:
            ready_count += 1

        ready_cell = ws.cell(row=row, column=col_ready_idx, value="Yes" if ready else "No")
        gaps_cell = ws.cell(row=row, column=col_gaps_idx, value="; ".join(remaining) if remaining else "")
        ready_cell.fill = READY_FILL if ready else NOT_READY_FILL
        gaps_cell.alignment = WRAP

        link_cell = ws.cell(row=row, column=col_link_idx, value="Open skills page")
        link_cell.hyperlink = f"{SKILLS_HTML}#{job_id}"
        link_cell.font = LINK_FONT

        updated += 1

    # ── document the new columns in the Readme sheet, if present ──────────
    if "Readme" in wb.sheetnames:
        readme = wb["Readme"]
        last_row = readme.max_row
        existing_text = " ".join(str(c.value or "") for r in readme.iter_rows() for c in r)
        if COL_READY not in existing_text:
            doc_rows = [
                (COL_READY, "Yes/No -- computed from skills_development_tracker.html progress. "
                             "'Yes' means every skill gap flagged for this job has been checked off."),
                (COL_GAPS, "Semicolon-separated list of skill titles still unchecked for this job. "
                           "Empty when Skill-Ready? = Yes."),
                (COL_LINK, "Hyperlink to this job's entry in skills_development_tracker.html for "
                           "learning resources. Regenerated by scripts/sync_skill_readiness.py."),
            ]
            r = last_row + 2
            readme.cell(row=r, column=1, value="Skills tracker columns (auto-synced):").font = Font(bold=True)
            r += 1
            for name, desc in doc_rows:
                readme.cell(row=r, column=1, value=name).font = Font(bold=True)
                readme.cell(row=r, column=2, value=desc).alignment = WRAP
                r += 1

    wb.save(TRACKER)
    print(f"Synced skill-readiness for {updated} jobs into {TRACKER.name} -- {ready_count} ready to apply.")


if __name__ == "__main__":
    main()
