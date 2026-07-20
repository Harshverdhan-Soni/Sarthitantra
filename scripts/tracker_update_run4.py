"""
Append JOB-016 through JOB-020 (Run 4 — 2026-07-01) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-016",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "NorthBay Solutions",
        "role":        "Generative AI Lead — Agentic AI",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/northbay-solutions/jobs/generative-ai-lead-agentic-ai",
        "score":       85,
        "rationale":   "Near-perfect agentic AI Lead match: multi-agent, LangChain, CrewAI, tool-calling, team mentoring. India remote. Salary 45-55 LPA. SAHAYAK-AI directly maps.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 85). Gap: AutoGPT/Semantic Kernel not explicit in profile but SAHAYAK-AI MCP architecture covers the orchestration patterns. Watch: consulting model, global timezones.",
        "resume":      "jobs/NorthBay_GenerativeAILead-AgenticAI/NorthBay_GenerativeAILead-AgenticAI_resume.docx",
        "cover":       "jobs/NorthBay_GenerativeAILead-AgenticAI/NorthBay_GenerativeAILead-AgenticAI_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-017",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "NorthBay Solutions",
        "role":        "Azure AI (Agentic) Specialist — Team Lead",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/northbay-solutions/jobs/azure-ai-agentic-specialist-team-lead",
        "score":       75,
        "rationale":   "Strong RAG pipeline, Knowledge Graphs, Agentic AI, govt client fit. India remote. Gap: Azure-specific services (Azure AI Search, Prompt Flow) vs GCP/OpenAI stack.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second top pick (score 75). Very recently posted (3 days). Cover letter honestly flags Azure-specific gap; transferable RAG/KG/agentic architecture is strong. Middle East govt client is direct experience match.",
        "resume":      "jobs/NorthBay_AzureAIAgenticSpecialist-TeamLead/NorthBay_AzureAIAgenticSpecialist-TeamLead_resume.docx",
        "cover":       "jobs/NorthBay_AzureAIAgenticSpecialist-TeamLead/NorthBay_AzureAIAgenticSpecialist-TeamLead_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-018",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Featherless AI",
        "role":        "AI Researcher — Distillation",
        "location":    "Worldwide",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/featherless-ai/jobs/ai-researcher-distillation",
        "score":       60,
        "rationale":   "AI research background fits; distillation/teacher-student pipelines are highly specialised and not explicit in profile; distributed training at scale is a gap.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 60 — below 70 threshold. Not tailored. Different role from JOB-002 (Multilingual). Could revisit if next run has no better options; PyTorch/model compression experience could be framed.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-019",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Wadhwani AI",
        "role":        "Machine Learning Scientist (ASR)",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/wadhwani-ai/jobs/machine-learning-scientist-asr",
        "score":       58,
        "rationale":   "Social-impact AI and public health domain alignment is excellent; however ASR/speech recognition is a specialisation gap — not in profile skills or projects.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 58 — below 70 threshold. Not tailored. Wadhwani AI mission (public health, NE India, nonprofit) is very appealing. Flag for future run if they post an NLP/RAG/GenAI role.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-020",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "ONE (OnePay)",
        "role":        "Senior Applied Scientist",
        "location":    "Bengaluru, India",
        "work_mode":   "Office (Bangalore)",
        "url":         "https://himalayas.app/companies/onepay/jobs/senior-applied-scientist",
        "score":       65,
        "rationale":   "ML/AI model development and LLM fit; fintech domain mismatch, Bangalore office (not remote preferred), Databricks not in profile stack.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 65 — below 70 threshold. Not tailored. Bangalore office conflicts with remote preference. Fintech (fraud, payments) is a domain gap. Could revisit if remote variant posted.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-016 through JOB-020 appended to {WB_PATH}")
