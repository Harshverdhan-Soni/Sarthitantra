"""
Daily run (2026-07-16) - append this run's 5 sourced roles to applications_tracker_NEW.xlsx. (Run 17)

  * Resolves the tracker via pathlib (BASE = this file's parent's parent), NOT a hardcoded path.
  * Dedupes on Job URL and Company+Role (safety net).
  * Assigns new Job IDs continuing from the highest JOB-### already in the sheet.

Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent          # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Cambium Learning Group",
        "role": "AI Architect",
        "location": "Remote (all countries, incl. India)", "work_mode": "Remote (Remote-First)",
        "url": "https://himalayas.app/companies/cambium-learning-group/jobs/ai-architect-2355120349",
        "score": 88,
        "rationale": "Top pick this run. Remote, open to all countries (India-eligible). Near-perfect map to profile "
                     "Sec 7/9: design end-to-end GenAI architectures with RAG, vector & hybrid search, grounding, "
                     "prompt orchestration, model evaluation; Python + LangChain/LangGraph + PyTorch; agentic "
                     "workflows/copilots; MLOps (registry, CI/CD, monitoring). Directly matches GCP-GraphRAG "
                     "(RAG + Neo4j KG + GNN, GCON 2026) and SAHAYAK-AI (multi-agent, RBAC, HITL). EdTech mission "
                     "(seen/valued/supported) aligns with public-good impact focus. 5+ yrs vs 12+ - clears easily.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Remote/all-countries, GenAI must-have met, no sensitive fields to apply. Azure "
                 "OpenAI/Vertex/Bedrock cloud-AI and cert preferences are 'nice-to-have' (upskill flag, not a block). "
                 "Apply-before Aug 28 2026. Note: AI tools prohibited during their interviews.",
        "resume": "jobs/CambiumLearning_AIArchitect/CambiumLearning_AIArchitect_resume.docx",
        "cover":  "jobs/CambiumLearning_AIArchitect/CambiumLearning_AIArchitect_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Addepto",
        "role": "AI Solution Architect",
        "location": "Remote (all countries, incl. India)", "work_mode": "Remote (B2B or contract of mandate)",
        "url": "https://himalayas.app/companies/addepto/jobs/ai-solution-architect",
        "score": 86,
        "rationale": "Second pick. Remote, open to all countries (India-eligible). Maps to 'AI Consultant' target "
                     "title (Sec 9): go-to AI expert, end-to-end AI architecture (data pipelines, model envs, "
                     "integrations, MLOps), framework/model selection, make-vs-buy, AI strategy/roadmaps/hackathons. "
                     "Strong ML + Generative AI + Agentic AI + PyTorch overlap with profile; SAHAYAK-AI (agents, "
                     "workflow automation) and GCP-GraphRAG fit well. Forbes top-10 AI consultancy. Slightly "
                     "consulting/industrial-leaning (predictive maintenance a plus) vs research; German a plus (not "
                     "required).",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Remote/all-countries, GenAI must-have met. B2B/contract-of-mandate cooperation and "
                 "occasional travel for onboarding/workshops - confirm acceptable. TensorFlow listed but PyTorch "
                 "accepted; SAP AI is a nice-to-have. Apply-before Sep 03 2026.",
        "resume": "jobs/Addepto_AISolutionArchitect/Addepto_AISolutionArchitect_resume.docx",
        "cover":  "jobs/Addepto_AISolutionArchitect/Addepto_AISolutionArchitect_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "TechBiz Global",
        "role": "Senior AI DevOps / LLMOps",
        "location": "Remote (multi-country, incl. India)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/techbiz-global/jobs/senior-ai-devops-llmops",
        "score": 68,
        "rationale": "Below threshold (not tailored). Remote incl. India, senior. LLMOps touches GenAI (LLM "
                     "deployment/serving) which overlaps profile Sec 7 (Docker/K8s, CI/CD, observability), but the "
                     "role is infrastructure/DevOps-centric - pipelines, serving, monitoring - rather than the "
                     "RAG/agentic build-and-research depth that is Harsh's core strength. AI/ML component present so "
                     "not a deal-breaker; just under the 70 bar.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (DevOps/LLMOps infra focus vs GenAI/RAG/agentic build strengths in profile Sec 7). "
                 "AI/ML present so not a strict deal-breaker. Below threshold - not tailored. Confirm full JD before "
                 "applying.",
        "resume": "", "cover": "", "fill": PINK,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Experior Financial Group",
        "role": "AI Engineer",
        "location": "Remote (multi-country, incl. India)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/experior-financial-group/jobs/ai-engineer-3257047063",
        "score": 66,
        "rationale": "Below threshold (not tailored). Remote incl. India. AI Engineer title fits the family, but this "
                     "is a financial-services generalist AI role with unclear GenAI/RAG/agentic depth (the profile's "
                     "Sec 7 core). AI/ML component present so not a deal-breaker; scope and domain fit are moderate, "
                     "placing it just under the 70 bar and behind the two tailored architect roles.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (financial-services generalist AI role; GenAI/RAG depth unclear vs profile Sec 7). "
                 "AI/ML present so not a strict deal-breaker. Below threshold - not tailored. Confirm full JD before "
                 "applying.",
        "resume": "", "cover": "", "fill": PINK,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Eqvilent",
        "role": "ML Engineer",
        "location": "Remote (incl. India)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/eqvilent/jobs/ml-engineer",
        "score": 62,
        "rationale": "Below threshold (not tailored). Remote incl. India. Quant-trading ML role - strong classical/"
                     "low-latency ML but no GenAI/RAG/agentic/multilingual-NLP component, which is Harsh's core per "
                     "Sec 7/9. AI/ML present so not a hard deal-breaker, but domain (trading signals/latency) and "
                     "focus diverge sharply from the applied-GenAI target profile. Not tailored.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (quant-trading ML domain, no GenAI/RAG/NLP overlap vs profile Sec 7). AI/ML present so "
                 "not a strict deal-breaker. Below threshold - not tailored.",
        "resume": "", "cover": "", "fill": PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

existing_urls, existing_pairs = set(), set()
max_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    jid = str(row[0]).strip() if row[0] else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
    company = (str(row[3]).strip().lower() if len(row) > 3 and row[3] else "")
    role    = (str(row[4]).strip().lower() if len(row) > 4 and row[4] else "")
    url     = (str(row[7]).strip().lower() if len(row) > 7 and row[7] else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

next_id = max_id + 1
appended, skipped = [], []

for job in jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        skipped.append(f'{job["company"]} - {job["role"]} (already in tracker)')
        continue

    job_id = f"JOB-{next_id:03d}"
    next_id += 1

    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    appended.append(f'{job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker updated:", WB_PATH)
for a in appended:
    print("  appended", a)
for s in skipped:
    print("  skipped ", s)
