"""
cancel_application.py - CANCEL an approved/queued application before a scheduled run
acts on it. Use this when a role is "Awaiting approval" (queued for pre-fill/apply) but
you have changed your mind and do NOT want the next run to pre-fill or apply to it.

It does NOT delete the job - the row stays in the tracker for your records, just marked
Cancelled so apply_approved.py / the scheduled run will skip it. Use delete_job.py if you
want the row gone entirely.

Usage (from the scripts/ folder):
    python cancel_application.py JOB-048
    python cancel_application.py JOB-048 --reason "Not interested after reading JD"
    python cancel_application.py JOB-048 JOB-091           (cancel several at once)

For each Job ID it:
    * Status         -> "Cancelled"
    * Outcome/notes  -> appends "Cancelled by user on <today>. <reason>"
    * Last updated   -> today
    * Row colour     -> grey (#D9D9D9)
    * apply_queue.json -> removes that job so the next run won't pre-fill/apply it
"""
import sys
import re
import json
import pathlib
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

import json as _json

BASE = pathlib.Path(__file__).parent.parent
QUEUE_PATH = BASE / "apply_queue.json"

def _tracker_path():
    at = BASE / "active_track.json"
    if at.exists():
        try:
            data = _json.loads(at.read_text(encoding="utf-8"))
            return str(BASE / data.get("tracker_file", "applications_tracker.xlsx"))
        except Exception:
            pass
    return str(BASE / "applications_tracker.xlsx")

WB_PATH = _tracker_path()

GREY = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

COL_ID, COL_COMPANY, COL_ROLE = 1, 4, 5
COL_STATUS, COL_NOTES, COL_OUTCOME, COL_UPDATED = 11, 13, 18, 19
TODAY = date.today().isoformat()


def parse_args(argv):
    job_ids, reason = [], None
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--reason":
            reason = argv[i + 1] if i + 1 < len(argv) else None
            i += 2
            continue
        m = re.match(r"(?i)job-?(\d+)", a)
        if m:
            job_ids.append("JOB-%03d" % int(m.group(1)))
        i += 1
    return job_ids, reason


def cancel_rows(ws, job_ids, reason):
    done = []
    for jid in job_ids:
        found = False
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, COL_ID).value).strip().upper() == jid:
                found = True
                ws.cell(r, COL_STATUS).value = "Cancelled"
                note = "Cancelled by user on " + TODAY + ("." if not reason else ". " + reason)
                cur = ws.cell(r, COL_OUTCOME).value or ""
                ws.cell(r, COL_OUTCOME).value = (cur + " | " + note) if cur else note
                ws.cell(r, COL_UPDATED).value = TODAY
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = GREY
                    ws.cell(r, c).alignment = WRAP
                ws.cell(r, COL_ID).font = BOLD
                done.append(jid + ": " + str(ws.cell(r, COL_COMPANY).value)
                            + " - " + str(ws.cell(r, COL_ROLE).value) + "  ->  Cancelled")
                break
        if not found:
            done.append(jid + ": NOT FOUND in tracker")
    return done


def prune_queue(job_ids):
    if not QUEUE_PATH.exists():
        return 0
    try:
        data = json.load(open(QUEUE_PATH))
    except Exception:
        return 0
    if not isinstance(data, list):
        return 0
    before = len(data)
    data = [j for j in data if str(j.get("job_id", "")).upper() not in set(job_ids)]
    json.dump(data, open(QUEUE_PATH, "w"), indent=2)
    return before - len(data)


def main():
    job_ids, reason = parse_args(sys.argv[1:])
    if not job_ids:
        try:
            entered = input("Enter the Job ID to CANCEL (e.g. JOB-048): ").strip()
        except EOFError:
            entered = ""
        m = re.match(r"(?i)job-?(\d+)", entered)
        if not m:
            print("No valid Job ID given. Nothing changed.")
            return
        job_ids = ["JOB-%03d" % int(m.group(1))]

    wb = load_workbook(WB_PATH)
    ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
    results = cancel_rows(ws, job_ids, reason)
    wb.save(WB_PATH)
    removed = prune_queue(job_ids)

    print("Tracker updated:", WB_PATH)
    for r in results:
        print("  ", r)
    print("  Removed from apply_queue.json:", removed, "entry(ies).")


if __name__ == "__main__":
    main()
