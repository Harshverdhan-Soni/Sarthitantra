"""
Daily run (2026-07-21) - Run 25.

Appends the 5 NEW roles sourced this run to applications_tracker_NEW.xlsx.
Tracker path resolved via pathlib (BASE = this file's parent's parent) - never hardcoded.
Dedupe is by Job URL and Job ID, so re-running is safe (already-logged rows are skipped).

Colour coding:
  GREEN  #D9EAD3 = Tailored
  YELLOW #FFF2CC = Scored / above threshold but not tailored
  PINK   #FFDACC = Below threshold or Skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import date

BASE = pathlib.Path(__file__).parent.parent
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK = PatternFill("solid", fgColor="FFDACC")
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()

new_jobs = [
    {
        "id": "JOB-121", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Saaf Finance",
        "role": "Forward-Deployed AI Engineer",
        "location": "India only (Remote)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/saaf-finance/jobs/forward-deployed-ai-engineer-8721758659",
        "score": 92,
        "rationale": "Top pick this run. LIVE (posted Jul 07 2026, apply before Sep 05 2026), India-only remote so no "
                     "sponsorship needed. Exceptional profile-Sec-7 match: build multi-step LLM agents with tool use, "
                     "structured output, memory and human-in-the-loop; RAG over documents; document extraction (OCR) "
                     "turning unstructured income/financial docs into structured data; evaluation frameworks and "
                     "data-flywheel feedback; strong production Python; AWS/containerisation; n8n workflow automation. "
                     "This is almost a description of Harsh's own systems - Finance SAHAYAK (multimodal RAG over "
                     "financial tables/charts/text, OCR + LlamaIndex) and SAHAYAK-AI (agentic, tool use, "
                     "human-in-the-loop, RBAC/audit governance). Fintech/regulated-data domain matches his governed "
                     "government-systems background. Banded mid-level, which he exceeds (12+ yrs).",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, applied-GenAI/agentic must-have met, no sensitive fields required "
                 "to apply. Tailored this run: resume leads with agentic LLM systems + RAG/document intelligence + "
                 "human-in-the-loop/evals + production Python/AWS; cover letter hooks on mortgage-lending AI, Finance "
                 "SAHAYAK and shipping to production. Note: aggregator (Himalayas) link stored; confirm Saaf Finance's "
                 "official careers / ATS posting as the apply source before submitting.",
        "resume": "jobs/SaafFinance_ForwardDeployedAIEngineer/SaafFinance_ForwardDeployedAIEngineer_resume.docx",
        "cover": "jobs/SaafFinance_ForwardDeployedAIEngineer/SaafFinance_ForwardDeployedAIEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-122", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "3Pillar Global",
        "role": "AI Lead Engineer (Prompt Engineering, RAG and LLM's)",
        "location": "India only (Remote)",
        "work_mode": "Remote (Contractor)",
        "url": "https://himalayas.app/companies/3pillar-global/jobs/ai-lead-engineer-prompt-engineering-rag-and-llm-s",
        "score": 90,
        "rationale": "Second tailored pick. LIVE (posted Jul 19 2026, apply before Sep 17 2026), India-only remote. "
                     "Clean match to profile Sec 7: strong prompt engineering with context injection, prompt "
                     "libraries, Bedrock/Claude integration and configuration-driven model selection for inference "
                     "cost; LLM concepts, information retrieval, recommendation and RAG; ML/NLP/deep learning in "
                     "Python (PyTorch, TensorFlow, scikit-learn); plus data-engineering to shape data for LLMs and "
                     "mentoring of junior AI/ML engineers - all directly in Harsh's toolkit. His GCP-GraphRAG and "
                     "Finance SAHAYAK cover the RAG/retrieval and cost/token-efficiency angle strongly, and the JD "
                     "explicitly values a desire to do AI research. Senior, 5+ yrs (he has 12+). Contractor "
                     "engagement. Minor gap: role names Amazon Bedrock specifically vs his Claude/GPT/Ollama delivery "
                     "- transferable and flagged honestly in the cover letter.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, GenAI/LLM must-have met, no sensitive fields required to apply. "
                 "Tailored this run: resume leads with prompt engineering + RAG/information retrieval + NLP/deep "
                 "learning + model-cost routing; cover letter hooks on Bedrock/Claude prompt integration and is "
                 "candid on Bedrock-specific depth vs his Claude/GPT/Ollama background. Note: aggregator (Himalayas) "
                 "link stored; confirm 3Pillar's official careers posting before submitting.",
        "resume": "jobs/3PillarGlobal_AILeadEngineer-PromptRAGLLM/3PillarGlobal_AILeadEngineer-PromptRAGLLM_resume.docx",
        "cover": "jobs/3PillarGlobal_AILeadEngineer-PromptRAGLLM/3PillarGlobal_AILeadEngineer-PromptRAGLLM_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-123", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "AiMi Technologies AB",
        "role": "Full Stack AI Engineer",
        "location": "India only (Remote)",
        "work_mode": "Remote (Contractor)",
        "url": "https://himalayas.app/companies/aimi-technologies-ab/jobs/full-stack-ai-engineer",
        "score": 78,
        "rationale": "Above threshold, NOT tailored (not top 2). India-only remote, contractor. AiMi Technologies is "
                     "a Swedish firm ('AB'), which aligns with Harsh's stated Europe/Sweden interest while still "
                     "being an India-remote posting (no relocation/sponsorship needed to apply). Full Stack AI "
                     "Engineer is a strong double-fit for his profile - AI/GenAI plus his Java/Spring Boot + ReactJS "
                     "full-stack delivery and Python. Held below the top 2 pending the detailed JD: full-stack-AI "
                     "roles vary widely in how much genuine LLM/RAG depth they carry vs. app plumbing, and this is a "
                     "contractor engagement with less research emphasis than his two tailored picks.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Provisional - India-only remote and AI/ML present; Swedish company is a nice long-term "
                 "signal for his EU/Sweden goal. Confirm full JD (LLM/RAG depth vs. general app work) and contractor "
                 "terms before tailoring. Surfaced for Harsh's decision - say the word and it can be tailored / "
                 "pre-filled. Aggregator (Himalayas) link stored; confirm AiMi's official careers page before "
                 "submitting.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "id": "JOB-124", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Nerdy",
        "role": "Staff Engineer (AI-Native)",
        "location": "India - Hyderabad (Remote / FT Contractor)",
        "work_mode": "Remote (Full-time Contractor)",
        "url": "https://himalayas.app/companies/nerdy/jobs/staff-engineer-ai-native-hyderabad-full-time-contractor",
        "score": 76,
        "rationale": "Above threshold, NOT tailored (not top 2). India (Hyderabad) staff-level AI-native engineering "
                     "role. Staff Engineer seniority suits his 12+ yrs and PhD-track depth, and an AI-native "
                     "build role should lean on his Python/GenAI/LLM strengths. Held below the top 2 for two "
                     "reasons: (1) the posting is a Hyderabad full-time-contractor arrangement, so degree of remote "
                     "vs. on-site and contract terms need confirming against his Remote / MP-preference target; and "
                     "(2) without the full JD it is unclear how much of the role is GenAI/RAG vs. general "
                     "AI-assisted product engineering. Solid, genuinely India-eligible, but a notch below the two "
                     "clean RAG/agentic matches.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Provisional - India (Hyderabad) and AI-native must-have plausibly met; confirm remote vs. "
                 "Hyderabad on-site and full-time-contractor terms, plus the detailed JD's GenAI depth, before "
                 "tailoring. Surfaced for Harsh's decision - can be tailored / pre-filled on request. Aggregator "
                 "(Himalayas) link stored; confirm Nerdy's official careers page before submitting.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "id": "JOB-125", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Hypersonix",
        "role": "Data Scientist_ML (India)",
        "location": "India only (Remote)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/hypersonix/jobs/data-scientist_ml-india",
        "score": 72,
        "rationale": "Above threshold, NOT tailored (not top 2). India-only remote, full-time. AI/ML must-have is met "
                     "and the ML/Python/deep-learning layer overlaps his Sec-7 keyword bank. Held lower and not "
                     "tailored because a 'Data Scientist_ML' role at a retail/commerce-AI company is likely centred "
                     "on classical/predictive ML and analytics rather than his GenAI / RAG / GraphRAG / "
                     "multilingual-NLP core - a genuine (if modest) domain lean away from his strongest, "
                     "differentiated work. No AI/ML deal-breaker; a reasonable India-remote option but not a top-tier "
                     "match this run.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes on India-remote + AI/ML present, but the DS/ML centre of gravity (predictive/analytics "
                 "ML) is a lean away from his GenAI/RAG strengths, so not auto-tailored. Detailed JD did not fully "
                 "load from the aggregator - confirm scope and Hypersonix's official careers posting before "
                 "tailoring/submitting. Surfaced for Harsh's decision.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications']

existing_urls = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}
existing_ids = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}

added = 0
for job in new_jobs:
    if job['url'] in existing_urls or job['id'] in existing_ids:
        print(f"SKIP (already tracked): {job['id']} {job['company']} - {job['role']}")
        continue
    r = ws.max_row + 1
    row = [
        job['id'], job['date'], job['source'], job['company'], job['role'],
        job['location'], job['work_mode'], job['url'], job['score'], job['rationale'],
        job['status'], job['dealbreaker'], job['notes'], job['resume'], job['cover'],
        '', '', '', TODAY, '', '', '',
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = job['fill']
        cell.alignment = WRAP
    added += 1
    print(f"ADDED {job['id']}: {job['company']} - {job['role']} (score {job['score']}, {job['status']})")

wb.save(WB_PATH)
print(f"\nSaved {WB_PATH} - {added} row(s) appended. Last row now {ws.max_row}.")
