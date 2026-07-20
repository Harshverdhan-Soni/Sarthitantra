"""
Daily run (2026-07-17) - Run 21b supplement.

Purpose:
  * Append the two remaining NEW roles for this run: Miratech (Senior GenAI Developer, TAILORED)
    and AHEAD (GenAI Engineer, Scored). Run 21 already added Kognitiv/Altisource/Cynet.
  * Upgrade existing rows that were only "Scored" in prior runs but are the highest-fit LIVE roles
    and now have tailored docs generated this run:
       - JOB-038 Luma Financial Technologies (AI/ML Engineer, LIVE, score 90) -> Tailored + doc paths.
       - JOB-012 Mactores (Generative AI Engineer, score 85) -> note tailored docs exist; posting EXPIRED,
         so status left as Scored (do not pre-fill an expired posting).

Resolves the tracker via pathlib (BASE = this file's parent's parent).

Color coding: GREEN=#D9EAD3 (Tailored), YELLOW=#FFF2CC (Scored/above-threshold not tailored).
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

new_jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app -> official SmartRecruiters)",
        "company": "Miratech",
        "role": "Senior GenAI Developer",
        "location": "India (All Cities, Remote)", "work_mode": "Remote",
        "url": "https://jobs.smartrecruiters.com/Miratech1/744000123670329-senior-genai-developer",
        "score": 82,
        "rationale": "Top NEW pick this run (LIVE official SmartRecruiters posting - canonical apply source). India "
                     "(All Cities) remote. Strong fit to profile Sec 7: build production RAG pipelines and agentic "
                     "workflows with LangChain/LangGraph; multi-agent architectures (Agentic RAG, Self-RAG, "
                     "Corrective RAG); advanced retrieval (hybrid, HyDE, chunking, semantic search); vector DBs "
                     "(PGVector/Pinecone/Weaviate); LLM integration (Claude/GPT-4o/Llama/Mistral); FastAPI/Docker. "
                     "Nice-to-haves Graph RAG and Human-in-the-Loop are Harsh's exact specialties (GCP-GraphRAG, "
                     "SAHAYAK-AI HITL). Gap: 8+ yrs cloud/AWS with Bedrock/Lambda hands-on (profile is Ollama/"
                     "on-prem + AWS-ready) and Ragas/LangSmith tooling - transferable, framed as growth.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India remote, GenAI/agentic must-have met, official live posting, no sensitive "
                 "fields to apply. AWS Bedrock/Lambda + IaC (Terraform) are the main gaps vs on-prem experience - "
                 "confirm comfort. Tailored this run (Graph RAG + HITL hooks).",
        "resume": "jobs/Miratech_SeniorGenAIDeveloper/Miratech_SeniorGenAIDeveloper_resume.docx",
        "cover":  "jobs/Miratech_SeniorGenAIDeveloper/Miratech_SeniorGenAIDeveloper_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "AHEAD (ThinkAhead)",
        "role": "GenAI Engineer",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/thinkahead/jobs/genai-engineer",
        "score": 80,
        "rationale": "Above threshold, not tailored (2nd-highest NEW role but full JD not confirmed live this run). "
                     "India-only remote; GenAI Engineer at an enterprise cloud/AI consultancy - GenAI must-have "
                     "present and title aligns with profile Sec 9 target titles. Strong likely fit (LLM/RAG/GenAI "
                     "build). Scored from listing metadata + category tags; the Himalayas apply link redirected on "
                     "fetch, so the official AHEAD careers posting must be confirmed before pre-fill.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (pending live-JD confirmation). India-only remote; GenAI must-have met. Confirm the "
                 "posting is live and find the official AHEAD/ThinkAhead careers page before any pre-fill. Above "
                 "threshold; not tailored this run (no confirmed JD to mirror without overstating).",
        "resume": "", "cover": "", "fill": YELLOW,
    },
]

# Upgrades to existing rows (matched by Job ID)
upgrades = {
    "JOB-038": {
        "status": "Tailored",
        "resume": "jobs/LumaFinancialTechnologies_AIMLEngineer/LumaFinancialTechnologies_AIMLEngineer_resume.docx",
        "cover":  "jobs/LumaFinancialTechnologies_AIMLEngineer/LumaFinancialTechnologies_AIMLEngineer_cover.docx",
        "note_append": ("[2026-07-17 Run21] Tailored resume + cover generated - highest-fit LIVE role available "
                        "(score 90; apply before 2026-09-02). Upgraded to Tailored. Ready for pre-fill/approval "
                        "on the official Luma careers/ATS posting."),
        "fill": GREEN,
    },
    "JOB-012": {
        "status": None,  # leave as-is (Scored); posting expired
        "resume": "jobs/Mactores_GenerativeAIEngineer/Mactores_GenerativeAIEngineer_resume.docx",
        "cover":  "jobs/Mactores_GenerativeAIEngineer/Mactores_GenerativeAIEngineer_cover.docx",
        "note_append": ("[2026-07-17 Run21] Tailored resume + cover generated (score 85). Himalayas posting "
                        "apply-before was 2026-07-08 (EXPIRED); NOT pre-filling. Reuse docs if reposted or a live "
                        "official Mactores careers posting is found."),
        "fill": None,  # do not recolor an expired row
    },
}

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

# --- indices (1-based): A JobID=1 ... H URL=8, I score=9, K status=11, M notes=13, N resume=14, O cover=15, S last=19
COL_STATUS, COL_NOTES, COL_RESUME, COL_COVER, COL_LAST = 11, 13, 14, 15, 19

existing_urls, existing_pairs = set(), set()
max_id = 0
id_to_row = {}
for r in range(2, ws.max_row + 1):
    jid = str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
        id_to_row[jid.upper()] = r
    company = (str(ws.cell(r, 4).value).strip().lower() if ws.cell(r, 4).value else "")
    role    = (str(ws.cell(r, 5).value).strip().lower() if ws.cell(r, 5).value else "")
    url     = (str(ws.cell(r, 8).value).strip().lower() if ws.cell(r, 8).value else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

# --- apply upgrades ---
for jid, up in upgrades.items():
    r = id_to_row.get(jid.upper())
    if not r:
        print(f"  upgrade skipped (row not found): {jid}")
        continue
    if up["status"]:
        ws.cell(r, COL_STATUS).value = up["status"]
    if up["resume"]:
        ws.cell(r, COL_RESUME).value = up["resume"]
    if up["cover"]:
        ws.cell(r, COL_COVER).value = up["cover"]
    prev = ws.cell(r, COL_NOTES).value or ""
    ws.cell(r, COL_NOTES).value = (str(prev).rstrip() + " " + up["note_append"]).strip()
    ws.cell(r, COL_LAST).value = TODAY
    if up["fill"]:
        for c in range(1, 20):
            ws.cell(r, c).fill = up["fill"]
    print(f"  upgraded {jid} -> status={up['status'] or '(unchanged)'}")

# --- append new roles ---
next_id = max_id + 1
for job in new_jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        print(f'  skipped (dup): {job["company"]} - {job["role"]}')
        continue
    job_id = f"JOB-{next_id:03d}"; next_id += 1
    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for c in range(1, len(row) + 1):
        cell = ws.cell(last_row, c)
        cell.fill = job["fill"]; cell.alignment = WRAP
        if c == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    print(f'  appended {job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker saved:", WB_PATH)
