"""
Daily run (2026-07-19) - Run 23.

Appends the 5 NEW roles sourced this run to applications_tracker_NEW.xlsx.
Tracker path resolved via pathlib (BASE = this file's parent's parent) - never hardcoded.

Colour coding:
  GREEN  #D9EAD3 = Tailored
  YELLOW #FFF2CC = Scored / above threshold but not tailored
  PINK   #FFDACC = Below threshold or Skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import date

BASE = pathlib.Path(__file__).parent.parent
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK = PatternFill("solid", fgColor="FFDACC")
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()

new_jobs = [
    {
        "id": "JOB-111", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Chess.com",
        "role": "Senior ML Engineer",
        "location": "Worldwide remote (AU, CA, DE, IN, UK, US listed; India eligible)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/chess-com/jobs/senior-ml-engineer",
        "score": 86,
        "rationale": "Top pick this run. LIVE (posted Jun 10 2026, apply before Aug 09 2026), 100% remote with India "
                     "in the eligible-country list, so no sponsorship needed. The JD maps closely to profile Sec 7: "
                     "own ML product features idea-to-production (modelling, deployment, experimentation, impact); "
                     "design/build/ship scalable ML systems, data pipelines, training workflows and production model "
                     "serving; shape MLOps practices; and - explicitly - 'drive innovation in Generative AI, "
                     "exploring how we can best use LLMs and agents'. Harsh's applied GenAI + agentic work "
                     "(SAHAYAK-AI, Finance SAHAYAK, GCP-GraphRAG) and full-stack production delivery fit both the ML "
                     "systems and the GenAI mandate; the technical-leadership/mentorship expectation matches his "
                     "faculty and standards-setting role. Requirement is 5+ yrs ML + 1+ yr AI engineering (he "
                     "exceeds). Gap: less emphasis on training large models from scratch and single-cloud depth vs "
                     "on-prem/containerised - transferable, stated honestly, not overstated.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Remote incl. India (verified on JD country list), GenAI must-have met, no sensitive "
                 "fields required to apply. Tailored this run: resume leads with product-ML ownership + GenAI/agents "
                 "+ MLOps; cover letter hooks on Generative-AI innovation and technical leadership, candid on the "
                 "single-cloud gap. Note: aggregator (Himalayas) link stored; confirm Chess.com's official careers "
                 "posting as the apply source before submitting.",
        "resume": "jobs/ChessCom_SeniorMLEngineer/ChessCom_SeniorMLEngineer_resume.docx",
        "cover": "jobs/ChessCom_SeniorMLEngineer/ChessCom_SeniorMLEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-112", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Workiva",
        "role": "Staff Machine Learning Engineer",
        "location": "Remote (AU, CA, IN, UK, US listed; India eligible)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/workiva/jobs/staff-machine-learning-engineer",
        "score": 84,
        "rationale": "Second tailored pick. LIVE (posted Jun 23 2026, apply before Aug 22 2026), remote with India "
                     "in the eligible-country list. Strong structural fit: architect and deliver ML solutions with "
                     "MLOps best practices, build tools/systems/automation for rapid ML dev, high availability and "
                     "observability, integrate Generative AI into products. Stack maps 1:1 to profile Sec 7 - "
                     "Python, Go, Java, GitHub, Docker, Kubernetes, CI/CD, AWS, data + deployment pipelines - and "
                     "Harsh's Java/Spring Boot + ReactJS + containerised delivery is a genuine differentiator here. "
                     "The GRC / financial-reporting / compliance domain aligns with his governed government-finance "
                     "systems (Finance SAHAYAK, DVDMS) and enterprise controls (RBAC, row-level security, MCP "
                     "governance, auditability). Requirement 4+ yrs ML (he exceeds); Staff-level leadership/mentoring "
                     "matches. Gaps: 24x7 on-call rotation and deep managed-AWS depth vs on-prem - flagged honestly "
                     "in the cover letter.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Remote incl. India (verified on JD country list), GenAI must-have met, no sensitive "
                 "fields required to apply. Tailored this run: resume leads with ML platform/MLOps + GenAI-in-product "
                 "+ Python/Go/Java/Docker/K8s and governance; cover letter hooks on compliance-grade reliability and "
                 "is candid about on-call and single-cloud depth. Note: US salary band shown ($163-261K) is "
                 "US-referenced; India comp will differ. Confirm Workiva's official careers posting before "
                 "submitting.",
        "resume": "jobs/Workiva_StaffMLEngineer/Workiva_StaffMLEngineer_resume.docx",
        "cover": "jobs/Workiva_StaffMLEngineer/Workiva_StaffMLEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-113", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "AI Technology Partners (AITP)",
        "role": "Principal AI Strategist & Solution Architect",
        "location": "Remote (AU, CA, DE, IN, UK, US listed; India eligible; US-based preferred)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/ai-technology-partners-inc/jobs/principal-ai-strategist-solution-architect",
        "score": 76,
        "rationale": "Above threshold, NOT tailored (not top 2). LIVE (apply before Jul 23 2026), India in the "
                     "eligible-country list. Enormous keyword overlap with profile Sec 7: generative AI, RAG "
                     "architectures, vector databases, prompt engineering, LLMOps/MLOps, agentic frameworks "
                     "(LangChain, LangGraph, LlamaIndex, CrewAI, AutoGen, Semantic Kernel) - AITP's whole stack is "
                     "Harsh's toolkit. Held below the top 2 because the role is a client-facing PRINCIPAL STRATEGIST "
                     "/ pre-sales + delivery-oversight leadership position (C-level advisory, workshops, proposals, "
                     "SoWs) rather than hands-on GenAI build, which is where Harsh is strongest; it also asks 10-15+ "
                     "yrs of enterprise consulting/solution-architecture specifically and states 'US-based candidates "
                     "preferred'. Strong on paper for the AI content, weaker on the consulting-exec profile it "
                     "actually wants.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Provisional - India listed as eligible BUT 'Remote, US-based candidates preferred' and "
                 "the role targets 10-15+ yrs of enterprise consulting / pre-sales solution architecture, which is a "
                 "different profile from Harsh's applied-research + build background (he has 12+ yrs total but not in "
                 "enterprise pre-sales consulting). AI/ML skills fit is excellent; role-type and location-preference "
                 "are the gaps. Surfaced for Harsh's decision rather than auto-tailored. Deadline is near "
                 "(Jul 23 2026) - say the word and it can be tailored/pre-filled immediately.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "id": "JOB-114", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Monalee",
        "role": "Senior Machine Learning Engineer",
        "location": "Worldwide remote (26 countries incl. India; India eligible)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/monalee/jobs/senior-machine-learning-engineer",
        "score": 63,
        "rationale": "BELOW THRESHOLD (70) - not tailored, not pre-filled. LIVE (posted Jun 23 2026, apply before "
                     "Aug 22 2026), remote with India in the 26-country eligible list, and the AI/ML must-have is "
                     "clearly present (deep learning, PyTorch/TensorFlow). Scored below threshold because the core "
                     "requirement is 3D COMPUTER VISION - object detection, segmentation, depth estimation, DINOv2 / "
                     "ViTs / SAM, plus geospatial data (photogrammetry, LiDAR, satellite imagery) for a solar-"
                     "proposal platform. That is a specialised CV/geospatial track outside Harsh's core strengths "
                     "(GenAI, RAG/GraphRAG, GNN, multilingual NLP); shared ground is only the general DL/PyTorch/"
                     "Python/Docker/cloud-deployment layer, and the JD wants 2+ yrs specifically in computer vision. "
                     "Honest fit rather than a stretch.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (JD wants 2+ yrs focused computer-vision experience plus geospatial/photogrammetry/"
                 "LiDAR, which Harsh does not have; his DL work is GenAI/NLP/GNN, not 3D CV). AI/ML must-have is met "
                 "and it is remote incl. India, so it is surfaced for Harsh's decision rather than skipped outright - "
                 "but it is a genuine domain mismatch, not a threshold-shaving case. Revisit only if he wants to "
                 "pivot toward CV/geospatial ML.",
        "resume": "", "cover": "", "fill": PINK,
    },
    {
        "id": "JOB-115", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Welo Global (Welo Data)",
        "role": "Generative AI Analyst | Hindi",
        "location": "India only (Remote)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/welo-global/jobs/generative-ai-analyst-hindi",
        "score": 30,
        "rationale": "DEAL-BREAKER / SKIPPED. India-only remote and thematically brushes Harsh's multilingual-NLP "
                     "interest (English->Hindi/Indic evaluation, aligned with AI4Bharat/BHASHINI themes), which is "
                     "why it surfaced. But it is an ENTRY-LEVEL linguistic annotation / translation-quality-review "
                     "role: reviewing two machine-generated Hindi translations and picking the better one, plus "
                     "video/data annotation and content review. There is no ML/AI engineering, modelling or system-"
                     "building component - it is human-in-the-loop data labeling. That triggers the profile Sec 9 "
                     "deal-breaker 'Roles with no AI/ML component', and the entry-level banding sits far below a 12+ "
                     "yr PhD-track engineer.",
        "status": "Skipped", "dealbreaker": "Yes",
        "notes": "Eligible: No - deal-breaker (no AI/ML engineering component; entry-level data annotation / "
                 "translation QA). India-only remote and multilingual theme is relevant, but the work is linguistic "
                 "review/labeling, not ML build, so per profile Sec 9 it is auto-skipped and not tailored or "
                 "pre-filled. Logged for the audit trail only.",
        "resume": "", "cover": "", "fill": PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications']

existing_urls = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}
existing_ids = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}

added = 0
for job in new_jobs:
    if job['url'] in existing_urls or job['id'] in existing_ids:
        print(f"SKIP (already tracked): {job['id']} {job['company']} - {job['role']}")
        continue
    r = ws.max_row + 1
    row = [
        job['id'], job['date'], job['source'], job['company'], job['role'],
        job['location'], job['work_mode'], job['url'], job['score'], job['rationale'],
        job['status'], job['dealbreaker'], job['notes'], job['resume'], job['cover'],
        '', '', '', TODAY, '', '', '',
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = job['fill']
        cell.alignment = WRAP
    added += 1
    print(f"ADDED {job['id']}: {job['company']} - {job['role']} (score {job['score']}, {job['status']})")

wb.save(WB_PATH)
print(f"\nSaved {WB_PATH} - {added} row(s) appended. Last row now {ws.max_row}.")
