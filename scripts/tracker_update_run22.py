"""
Daily run (2026-07-18) - Run 22.

Appends the 5 NEW roles sourced this run to applications_tracker_NEW.xlsx.
Tracker path resolved via pathlib (BASE = this file's parent's parent) - never hardcoded.

Colour coding:
  GREEN  #D9EAD3 = Tailored
  YELLOW #FFF2CC = Scored / above threshold but not tailored
  PINK   #FFDACC = Below threshold or Skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import date

BASE = pathlib.Path(__file__).parent.parent
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK = PatternFill("solid", fgColor="FFDACC")
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()

new_jobs = [
    {
        "id": "JOB-106", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Nagarro",
        "role": "Principal Engineer, Python Fullstack (React + GenAI)",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/nagarro/jobs/principal-engineer-python-fullstack-react-genai",
        "score": 90,
        "rationale": "Top pick this run and the best structural fit sourced to date. LIVE (posted Jun 29 2026, apply "
                     "before Aug 28 2026), India-only remote. The JD is Harsh's exact dual profile in one role: 11+ "
                     "yrs total (he has 12+), deep Python AND React, plus proven delivery of production-grade "
                     "Generative AI at scale. Required skills map 1:1 to profile Sec 7 - LLM integration patterns, "
                     "RAG systems, agent-based orchestration over tools/APIs/structured data, AI-driven UX "
                     "(conversational interfaces, copilots, dashboards), prompt design, evaluation and "
                     "observability. 'Strong grasp of security, privacy and governance for enterprise AI' matches "
                     "SAHAYAK-AI (RBAC, row-level security, MCP governance, HITL). Full-stack national-scale "
                     "delivery (HMIS 93->132+ facilities on Spring Boot/ReactJS/PostgreSQL/Docker) evidences the "
                     "principal-level scope. JD even names Copilot/Claude Code as delivery tools, which he uses "
                     "daily. Gap: hands-on AWS/Azure/GCP at scale (his production stack is on-prem/Ollama + "
                     "containerised) - transferable, stated honestly in the cover letter, not overstated.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, GenAI must-have met, official live posting via Nagarro on "
                 "Himalayas, no sensitive fields required to apply. Only gap is depth of hands-on public-cloud "
                 "(AWS/Azure/GCP) architecture vs on-prem experience - flagged, not fabricated. Tailored this run: "
                 "resume leads with GenAI architecture + Python/React fullstack; cover letter hooks on enterprise "
                 "readiness and responsible AI. Note: canonical apply source should be confirmed on Nagarro's own "
                 "careers site before submitting.",
        "resume": "jobs/Nagarro_PrincipalEngineerPythonFullstackReactGenAI/Nagarro_PrincipalEngineerPythonFullstackReactGenAI_resume.docx",
        "cover": "jobs/Nagarro_PrincipalEngineerPythonFullstackReactGenAI/Nagarro_PrincipalEngineerPythonFullstackReactGenAI_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-107", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "GR8 Tech",
        "role": "Senior Artificial Intelligence Specialist (Senior AI / GenAI Engineer)",
        "location": "Worldwide remote (open to candidates from all countries; India eligible)",
        "work_mode": "Remote",
        "url": "https://himalayas.app/companies/gr8-tech/jobs/senior-artificial-intelligence-specialist-7831435069",
        "score": 87,
        "rationale": "Second tailored pick. LIVE but CLOSING SOON (apply before Jul 22 2026). Worldwide remote so "
                     "India-eligible with no sponsorship needed. Requirements read like profile Sec 7: production "
                     "LLM systems owned design-to-support, RAG architectures (ingestion/embeddings/retrieval/"
                     "generation), chatbots and assistants with predictable behaviour, tool/function calling and "
                     "agent-style patterns, and - notably - MCP or similar context/tool-serving systems, which is a "
                     "named Harsh specialty (SAHAYAK-AI MCP governance). Stack overlap is high: Python, SQL, "
                     "LangChain/LangGraph, Anthropic/OpenAI/open-source models, pgvector, Docker, Git, CI/CD. "
                     "Explicitly 'not research, not prompt-only', which suits his applied + delivery mix. "
                     "Nice-to-haves (multi-agent, MLOps, production incident ownership) are all present. Gap: "
                     "preferred AWS depth (Bedrock/ECS/EKS/Lambda/OpenSearch) vs on-prem Ollama + containers - "
                     "stated honestly in the cover letter.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Worldwide remote, GenAI must-have met, no sensitive data required to apply. "
                 "TWO FLAGS FOR HARSH: (1) DEADLINE - apply before Jul 22 2026, so this needs submitting within "
                 "days; (2) DOMAIN - GR8 Tech is a B2B iGaming / online sports betting and casino platform "
                 "provider. That is not a deal-breaker in profile Sec 9, but it is a sector choice worth a "
                 "deliberate decision before applying, so it is surfaced rather than assumed. Tailored this run: "
                 "resume leads with production LLM ownership, RAG and MCP/tool-serving; cover letter hooks on "
                 "failure modes and graceful degradation and is candid about the AWS gap.",
        "resume": "jobs/GR8Tech_SeniorAISpecialist/GR8Tech_SeniorAISpecialist_resume.docx",
        "cover": "jobs/GR8Tech_SeniorAISpecialist/GR8Tech_SeniorAISpecialist_cover.docx",
        "fill": GREEN,
    },
    {
        "id": "JOB-108", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Fusemachines",
        "role": "Applied AI Engineer (Automation)",
        "location": "India only (Remote)", "work_mode": "Remote (Contractor)",
        "url": "https://himalayas.app/companies/fusemachines/jobs/applied-ai-engineer-automation-8215597112",
        "score": 82,
        "rationale": "Above threshold, NOT tailored - see deadline note. India-only remote. Content fit is strong: "
                     "agentic workflows with LangChain/LangGraph and tool calling, RAG pipelines with vector DBs "
                     "(Pinecone/Elasticsearch/pgvector), workflow automation on n8n/Make/Zapier (n8n and Make are "
                     "both in profile Sec 7), Python/FastAPI services, Docker and cloud deployment, plus preferred "
                     "knowledge graphs and AI-governance/PII/auditability - all core Harsh material (SAHAYAK-AI, "
                     "Finance SAHAYAK, GCP-GraphRAG). Fusemachines' democratise-AI and AI-education mission also "
                     "aligns with his PG-DAI/PGDAC faculty work. Held below the top 2 mainly because it is a "
                     "CONTRACTOR engagement pitched at 3-8 yrs (mid-to-senior), which sits under his 12+ yrs and "
                     "his 15LPA expectation, and because the window is effectively closed.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes on skills, but URGENT/EXPIRING - apply-before date is Jul 19 2026, i.e. tomorrow. "
                 "Not tailored this run because a same-day contractor application at below-target seniority is a "
                 "judgement call for Harsh, not one to make automatically. If he wants it, say so and it can be "
                 "tailored and pre-filled immediately; otherwise it will lapse. Also note: contractor terms and "
                 "mid-level banding vs 12+ yrs / 15LPA expectation.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "id": "JOB-109", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "JLL",
        "role": "Principal Architect, AI Solutions",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/jll/jobs/principal-architect-ai-solutions-424435491",
        "score": 76,
        "rationale": "Above threshold, not top 2. India-only remote and the AI/ML must-have is present (principal-"
                     "level AI solution architecture). Title maps to profile Sec 9 target titles (AI Consultant / "
                     "architect-level) and to SAHAYAK-AI architecture, MCP governance and multi-agent design; "
                     "principal seniority suits 12+ yrs. Held back from tailoring because the role leans "
                     "architecture-and-advisory over hands-on GenAI build, which is where Harsh is strongest, and "
                     "because the full JD could not be confirmed live this run - scored from listing metadata and "
                     "skill tags rather than the complete description. Commercial real-estate domain is new ground.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (provisional). India-only remote, AI must-have met. Caveat: scored from aggregator "
                 "listing metadata only - the full JD was not confirmed live, so required years and any mandatory "
                 "cloud/certification bar are unverified. Confirm the official JLL careers posting before any "
                 "pre-fill. Domain (commercial real estate) is unfamiliar but not disqualifying.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "id": "JOB-110", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "LatentView Analytics",
        "role": "Associate Director, Head of Technology",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/latentview-analytics/jobs/associate-director-head-of-technology",
        "score": 66,
        "rationale": "BELOW THRESHOLD (70) - not tailored, not pre-filled. LIVE (apply before Sep 07 2026) and "
                     "India-only remote, and the AI content is genuinely relevant: hands-on LLMs, generative-AI "
                     "frameworks and agentic architectures, productionising AI/analytics at client scale. However "
                     "the JD states a hard bar of 15+ years of progressive technology experience plus 5+ years in a "
                     "senior ENGINEERING LEADERSHIP role owning multi-disciplinary teams. Harsh has 12+ years and "
                     "his leadership is project-coordination and faculty/mentoring rather than heading an "
                     "engineering function, so this is a stated minimum he does not currently meet. Scored honestly "
                     "on that gap rather than stretched to clear the threshold.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (stated minimum 15+ yrs total and 5+ yrs senior engineering leadership; Harsh has "
                 "12+ yrs and no head-of-function leadership tenure). Deep hands-on expertise in a major cloud "
                 "platform is a second gap. Not a deal-breaker role - the AI scope fits well - so it is surfaced "
                 "for Harsh's decision rather than skipped outright. Worth revisiting as a stretch application "
                 "only if he wants to test the director track; posting stays open to Sep 07 2026.",
        "resume": "", "cover": "", "fill": PINK,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications']

existing_urls = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}
existing_ids = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}

added = 0
for job in new_jobs:
    if job['url'] in existing_urls or job['id'] in existing_ids:
        print(f"SKIP (already tracked): {job['id']} {job['company']} - {job['role']}")
        continue
    r = ws.max_row + 1
    row = [
        job['id'], job['date'], job['source'], job['company'], job['role'],
        job['location'], job['work_mode'], job['url'], job['score'], job['rationale'],
        job['status'], job['dealbreaker'], job['notes'], job['resume'], job['cover'],
        '', '', '', TODAY, '', '', '',
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = job['fill']
        cell.alignment = WRAP
    added += 1
    print(f"ADDED {job['id']}: {job['company']} - {job['role']} (score {job['score']}, {job['status']})")

wb.save(WB_PATH)
print(f"\nSaved {WB_PATH} - {added} row(s) appended. Last row now {ws.max_row}.")
