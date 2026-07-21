"""
Daily run (2026-07-21) - Run 25b (supplement).

Hypersonix Data Scientist_ML was a genuine duplicate (already tracked as JOB-040),
so this adds one replacement India-only role to keep the run at 5 net-new listings.
Path resolved via pathlib; dedupe by URL/ID so re-running is safe.
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import date

BASE = pathlib.Path(__file__).parent.parent
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

YELLOW = PatternFill("solid", fgColor="FFF2CC")
WRAP = Alignment(wrap_text=True, vertical="top")
TODAY = date.today().isoformat()

new_jobs = [
    {
        "id": "JOB-125", "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Tech Holding",
        "role": "Senior FullStack Engineer - Gen AI (Contract)",
        "location": "India only (Remote)",
        "work_mode": "Remote (Contractor)",
        "url": "https://himalayas.app/companies/tech-holding/jobs/senior-fullstack-engineer-gen-ai-_-contract",
        "score": 75,
        "rationale": "Above threshold, NOT tailored (not top 2). India-only remote, contractor. Replacement for the "
                     "Hypersonix listing, which turned out to already be tracked (JOB-040). Strong double-fit: a "
                     "Senior Full-Stack Gen AI role maps onto Harsh's two strengths at once - GenAI/LLM work plus his "
                     "Java/Spring Boot + ReactJS + Python full-stack delivery (the JD skills list flags React). Held "
                     "below the top 2 because it is a contractor engagement and, like most full-stack-AI postings, "
                     "the balance of genuine LLM/RAG depth vs. general application engineering needs confirming from "
                     "the full JD before investing in a tailored resume.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Provisional - India-only remote and GenAI present with a clear full-stack overlap; confirm "
                 "the detailed JD (LLM/RAG depth vs. app work) and contractor terms before tailoring. Surfaced for "
                 "Harsh's decision - can be tailored / pre-filled on request. Aggregator (Himalayas) link stored; "
                 "confirm Tech Holding's official careers page before submitting.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications']

existing_urls = {ws.cell(r, 8).value for r in range(2, ws.max_row + 1)}
existing_ids = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}

added = 0
for job in new_jobs:
    if job['url'] in existing_urls or job['id'] in existing_ids:
        print(f"SKIP (already tracked): {job['id']} {job['company']} - {job['role']}")
        continue
    r = ws.max_row + 1
    row = [
        job['id'], job['date'], job['source'], job['company'], job['role'],
        job['location'], job['work_mode'], job['url'], job['score'], job['rationale'],
        job['status'], job['dealbreaker'], job['notes'], job['resume'], job['cover'],
        '', '', '', TODAY, '', '', '',
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = job['fill']
        cell.alignment = WRAP
    added += 1
    print(f"ADDED {job['id']}: {job['company']} - {job['role']} (score {job['score']}, {job['status']})")

wb.save(WB_PATH)
print(f"\nSaved {WB_PATH} - {added} row(s) appended. Last row now {ws.max_row}.")
