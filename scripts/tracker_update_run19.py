"""
Daily run (2026-07-16) - append this run's 5 sourced roles to applications_tracker_NEW.xlsx. (Run 19)

  * Resolves the tracker via pathlib (BASE = this file's parent's parent), NOT a hardcoded path.
  * Dedupes on Job URL and Company+Role (safety net).
  * Assigns new Job IDs continuing from the highest JOB-### already in the sheet.

Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent          # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Valent Procure",
        "role": "Founding Applied AI Engineer",
        "location": "India only (Remote)", "work_mode": "Remote (US ET overlap for part of day)",
        "url": "https://himalayas.app/companies/valent-procure/jobs/founding-applied-ai-engineer-india",
        "score": 86,
        "rationale": "Top pick this run. India-only remote, live (apply-before Aug 23 2026). Near-perfect fit to profile "
                     "Sec 7/9: LLM app dev, RAG & retrieval, document intelligence/extraction (OCR), agentic task "
                     "execution, and evaluation pipelines to reduce hallucination with auditable, human-in-the-loop "
                     "outputs - maps directly to Finance SAHAYAK (multimodal RAG+OCR), GCP-GraphRAG (provenance-aware, "
                     "token-efficient) and SAHAYAK-AI (multi-agent, HITL). 3+ yrs vs 12+. Founding role rewards full-"
                     "stack range (Python/TypeScript/React/Postgres) which Harsh has.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, applied-AI must-have met, no sensitive fields to apply. Minor "
                 "friction: asks for part-day overlap with US Eastern Time - confirm Harsh can accommodate. Apply "
                 "before Aug 23 2026.",
        "resume": "jobs/ValentProcure_FoundingAppliedAIEngineer/ValentProcure_FoundingAppliedAIEngineer_resume.docx",
        "cover":  "jobs/ValentProcure_FoundingAppliedAIEngineer/ValentProcure_FoundingAppliedAIEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Sezzle",
        "role": "AI Engineer II",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/sezzle/jobs/ai-engineer-ii-2362602988",
        "score": 82,
        "rationale": "Second pick. India-only remote, live (apply-before Aug 30 2026), salary $3,333-6,000/mo. Applied "
                     "AI team building AI-powered platforms from scratch, intelligent automation/orchestration, agentic "
                     "systems, RAG, NLP (classification/NER/semantic search), transformers/embeddings + vector DBs - "
                     "strong overlap with profile Sec 7 and SAHAYAK-AI/GCP-GraphRAG. Requires hands-on Claude/LLM use "
                     "(Harsh uses Claude daily) and full-stack Python/SQL/React + Docker/K8s. Mid-level 3-5 yrs; Harsh "
                     "well above. dbt/Golang are minor nice-to-have gaps.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. India-only remote, GenAI/agentic must-have met, Claude/LLM experience required and met. "
                 "Golang is 'a plus' (willing to learn) - not a blocker. Apply before Aug 30 2026.",
        "resume": "jobs/Sezzle_AIEngineerII/Sezzle_AIEngineerII_resume.docx",
        "cover":  "jobs/Sezzle_AIEngineerII/Sezzle_AIEngineerII_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Jeeves",
        "role": "Senior AI Engineer",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/jeeves/jobs/senior-ai-engineer-7141998569",
        "score": 84,
        "rationale": "High-fit role but POSTING EXPIRED (apply-before Jun 16 2026; today 2026-07-16). India-only remote. "
                     "Production LLM/RAG pipelines, prompt engineering, chain orchestration (LangChain/LlamaIndex), "
                     "vector search (Pinecone/Weaviate/pgvector), ML serving, human-in-the-loop for financial "
                     "decisions - excellent match to profile Sec 7 and SAHAYAK-AI/GCP-GraphRAG. Would score ~84 if "
                     "live. Not tailored because the listing has closed; surfaced for approval only.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (posting expired Jun 16 2026). Strong fit if a live re-post appears - check Jeeves "
                 "careers page (tryjeeves.com) for a current Senior AI Engineer opening before applying. Not tailored.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Lingaro",
        "role": "Regular Lead Data Scientist Gen AI",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/lingaro/jobs/regular-lead-data-scientist-gen-ai",
        "score": 80,
        "rationale": "Strong GenAI fit but POSTING EXPIRED (apply-before Jun 26 2026). India-only remote. End-to-end "
                     "Talk-to-Data GenAI apps, RAG pipelines with vector DBs/rerankers/eval frameworks, fine-tuning "
                     "(LoRA/QLoRA/SFT), multi-agent workflows and MCP connectors - maps cleanly to profile Sec 7 "
                     "(RAG/GraphRAG, LangChain/LangGraph/LlamaIndex, PyTorch, MCP, NLP). PhD-level GenAI team is a good "
                     "cultural match. Not tailored because listing has closed; surfaced for approval only.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: No (posting expired Jun 26 2026). Very good fit - check Lingaro careers (lingarogroup.com) "
                 "for a current Data Scientist Gen AI / ML-AI Engineer opening in India before applying. Not tailored.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Jitterbit",
        "role": "Senior AI Engineer",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/jitterbit/jobs/senior-ai-engineer-remote-9111336103",
        "score": 78,
        "rationale": "Above threshold, not top 2. India-only remote, live (apply-before Aug 30 2026). LLM, RAG, "
                     "LangChain/LlamaIndex, Azure AI/AWS Bedrock/OpenAI, Python/Java, Docker/K8s - solid overlap with "
                     "profile Sec 7. Leans heavily to backend/iPaaS infra and multi-tenant SaaS at 10+ yrs; Harsh has "
                     "12+ yrs total incl. Java/Spring Boot at scale, so meets the bar, but the role is more "
                     "platform/integration than the RAG/agentic research build that is Harsh's core strength.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (India-only remote, GenAI must-have met; 10+ yrs met via 12+ yrs incl. large-scale "
                 "Java/Spring Boot SaaS). Infra/iPaaS-heavy vs research focus; behind the two tailored roles. Apply "
                 "before Aug 30 2026. Not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

existing_urls, existing_pairs = set(), set()
max_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    jid = str(row[0]).strip() if row[0] else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
    company = (str(row[3]).strip().lower() if len(row) > 3 and row[3] else "")
    role    = (str(row[4]).strip().lower() if len(row) > 4 and row[4] else "")
    url     = (str(row[7]).strip().lower() if len(row) > 7 and row[7] else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

next_id = max_id + 1
appended, skipped = [], []

for job in jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        skipped.append(f'{job["company"]} - {job["role"]} (already in tracker)')
        continue

    job_id = f"JOB-{next_id:03d}"
    next_id += 1

    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    appended.append(f'{job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker updated:", WB_PATH)
for a in appended:
    print("  appended", a)
for s in skipped:
    print("  skipped ", s)
