"""
Daily run (2026-07-16) - append this run's 5 sourced roles to applications_tracker_NEW.xlsx. (Run 18)

  * Resolves the tracker via pathlib (BASE = this file's parent's parent), NOT a hardcoded path.
  * Dedupes on Job URL and Company+Role (safety net).
  * Assigns new Job IDs continuing from the highest JOB-### already in the sheet.

Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent          # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

jobs = [
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Newel Health",
        "role": "AI/ML Engineer",
        "location": "Remote (multi-country, incl. India)", "work_mode": "Remote (Europe-based preferred)",
        "url": "https://himalayas.app/companies/newel-health/jobs/ai-ml-engineer",
        "score": 84,
        "rationale": "Top pick this run. Remote, India explicitly listed as an eligible country. Strong nice-to-have "
                     "alignment: healthcare AI (SaMD/DTx digital therapies) + Explainable AI, reproducibility, bias "
                     "mitigation and privacy-preserving ML - all core to Harsh's PhD focus and government health-"
                     "informatics work (HMIS, 132+ facilities). Supervised/unsupervised ML, scalable pipelines, "
                     "PyTorch, MLOps and generative-AI evaluation map cleanly to profile Sec 7. 4+ yrs required vs "
                     "12+. Gaps (clinical signal processing, formal clinical validation) are learnable, not blockers.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Remote incl. India, GenAI/applied-AI must-have met, no sensitive fields to apply. "
                 "Europe-based 'preferred' (not required); confirm timezone overlap. Apply-before Aug 27 2026.",
        "resume": "jobs/NewelHealth_AIMLEngineer/NewelHealth_AIMLEngineer_resume.docx",
        "cover":  "jobs/NewelHealth_AIMLEngineer/NewelHealth_AIMLEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Shipwell",
        "role": "Machine Learning Engineer",
        "location": "Remote (all countries, incl. India)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/shipwell/jobs/machine-learning-engineer",
        "score": 80,
        "rationale": "Second pick. Remote, open to candidates from all countries (India-eligible). Full ML product "
                     "lifecycle + agentic tooling/pipelines (LangChain, LangGraph, LangSmith) + GenAI in logistics - "
                     "directly hits profile Sec 7 (agentic AI, LangChain/LangGraph, Python, Docker/K8s, Git, SQL/"
                     "Postgres). Maps to SAHAYAK-AI (multi-agent) and GCP-GraphRAG. Data-engineering-heavy (ETL, dbt, "
                     "data integrity) leans more infra than research; dbt is a specific tool gap. Senior, $130-163K.",
        "status": "Tailored", "dealbreaker": "No",
        "notes": "Eligible: Yes. Remote/all-countries, GenAI/agentic must-have met. dbt and deep AWS data-warehouse "
                 "experience are nice-to-have gaps (upskill flag, not a hard block). Apply-before Aug 30 2026.",
        "resume": "jobs/Shipwell_MachineLearningEngineer/Shipwell_MachineLearningEngineer_resume.docx",
        "cover":  "jobs/Shipwell_MachineLearningEngineer/Shipwell_MachineLearningEngineer_cover.docx",
        "fill": GREEN,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "ImagineArt",
        "role": "Principal AI Engineer (LLM Agents & Orchestration)",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/imagineart/jobs/principal-ai-engineer-llm-agents-orchestration-7413998027",
        "score": 79,
        "rationale": "Above threshold, not top 2. India-only remote; title maps to GenAI Engineer / Applied AI target. "
                     "LLM agents & orchestration = direct agentic-AI nice-to-have (SAHAYAK-AI, multi-agent). Scored "
                     "just below the two tailored roles because the JD page was not retrievable (redirected to the "
                     "generic listing - likely stale/expired), so Principal-level required skills could not be "
                     "verified against profile. Not tailored pending JD confirmation.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (India-only remote, agentic-AI must-have met by title) BUT JD page not retrievable "
                 "(redirected to /jobs; possibly expired). Verify live listing and Principal-level requirements "
                 "before applying. Not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Turing",
        "role": "AI Engineering Lead",
        "location": "India & US only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/turing/jobs/ai-engineering-lead",
        "score": 78,
        "rationale": "Above threshold, not top 2. IN/US remote; AI Engineering Lead maps to Senior AI/ML Engineer "
                     "target and LLM/GenAI delivery. Leadership + agentic/LLM scope fit profile seniority (12+ yrs, "
                     "PhD in progress). Scored below the tailored roles because the JD page redirected to the generic "
                     "listing (likely stale/expired) and specific requirements could not be verified. Not tailored.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (IN/US remote, GenAI must-have met by title) BUT JD page not retrievable (redirected "
                 "to /jobs; possibly expired). Confirm live listing before applying. Not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
    {
        "date": TODAY, "source": "Web Search (Himalayas.app)",
        "company": "Fello",
        "role": "AI Applied Engineer (GTM AI Team)",
        "location": "India only (Remote)", "work_mode": "Remote",
        "url": "https://himalayas.app/companies/fello/jobs/ai-applied-engineer-gtm-ai-team",
        "score": 74,
        "rationale": "Above threshold, not top 2. India-only remote; Applied AI Engineer title fits the family and "
                     "applies GenAI/LLM to go-to-market workflows - overlaps profile Sec 7 (GenAI, agentic, prompt "
                     "engineering, Python). GTM/sales-ops domain focus is narrower than Harsh's research + public-"
                     "sector strengths, and the JD was only available at listing level, so scored solidly above the "
                     "bar but behind the two tailored, fully-documented roles.",
        "status": "Scored", "dealbreaker": "No",
        "notes": "Eligible: Yes (India-only remote, GenAI must-have met). GTM/sales-ops domain fit is moderate; full "
                 "JD not fetched (listing-level only). Above threshold - not tailored this run.",
        "resume": "", "cover": "", "fill": YELLOW,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

existing_urls, existing_pairs = set(), set()
max_id = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or all(c is None for c in row):
        continue
    jid = str(row[0]).strip() if row[0] else ""
    m = re.match(r"JOB-(\d+)", jid, re.I)
    if m:
        max_id = max(max_id, int(m.group(1)))
    company = (str(row[3]).strip().lower() if len(row) > 3 and row[3] else "")
    role    = (str(row[4]).strip().lower() if len(row) > 4 and row[4] else "")
    url     = (str(row[7]).strip().lower() if len(row) > 7 and row[7] else "")
    if url:
        existing_urls.add(url)
    if company and role:
        existing_pairs.add((company, role))

next_id = max_id + 1
appended, skipped = [], []

for job in jobs:
    key_url  = job["url"].strip().lower()
    key_pair = (job["company"].strip().lower(), job["role"].strip().lower())
    if key_url in existing_urls or key_pair in existing_pairs:
        skipped.append(f'{job["company"]} - {job["role"]} (already in tracker)')
        continue

    job_id = f"JOB-{next_id:03d}"
    next_id += 1

    row = [
        job_id, job["date"], job["source"], job["company"], job["role"],
        job["location"], job["work_mode"], job["url"], job["score"], job["rationale"],
        job["status"], job["dealbreaker"], job["notes"], job["resume"], job["cover"],
        "", "", "", TODAY,
    ]
    ws.append(row)
    last_row = ws.max_row
    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD
    existing_urls.add(key_url); existing_pairs.add(key_pair)
    appended.append(f'{job_id}: {job["company"]} - {job["role"]} ({job["status"]}, score {job["score"]})')

wb.save(WB_PATH)
print("Tracker updated:", WB_PATH)
for a in appended:
    print("  appended", a)
for s in skipped:
    print("  skipped ", s)
