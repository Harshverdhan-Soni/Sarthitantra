"""
Append JOB-021 through JOB-025 (Run 5 — 2026-07-02) to applications_tracker_NEW.xlsx.
Color coding:
  GREEN  (#D9EAD3) = Tailored
  YELLOW (#FFF2CC) = Scored / above threshold, not tailored
  PINK   (#FFDACC) = Below threshold or skipped
"""
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

BASE    = pathlib.Path(__file__).parent.parent  # JobFinder root
WB_PATH = str(BASE / 'applications_tracker_NEW.xlsx')

GREEN  = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
PINK   = PatternFill("solid", fgColor="FFDACC")
BOLD   = Font(bold=True)
WRAP   = Alignment(wrap_text=True, vertical="top")
TODAY  = date.today().isoformat()

# Columns in tracker:
# Job ID | Date sourced | Source | Company | Role | Location | Work mode |
# Job URL | Fit score | Score rationale | Status | Deal-breaker? |
# Missing requirements / notes | Resume file | Cover letter file |
# Date submitted | Follow-up date | Outcome | Last updated

jobs = [
    {
        "job_id":      "JOB-021",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "SS&C Technologies",
        "role":        "GenAI Developer / Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/ss-c-technologies/jobs/genai-developer-engineer-2874542653",
        "score":       80,
        "rationale":   "Strong GenAI/LLM/RAG keyword match (prompt engineering, Azure OpenAI, Pinecone/Faiss/Weaviate) "
                        "directly hits profile's core GenAI stack; role asks for only 2-3 yrs GenAI exposure vs 12+ "
                        "yrs profile depth — solid but not the strongest fit of this batch. Gap: Pinecone/Weaviate "
                        "vector-DB stack not explicit in profile (Neo4j-based RAG instead).",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 80 — above threshold but not top 2 this run. 5000+ employee fintech/healthtech; "
                        "very recently posted (14 hrs). Could revisit if next run has fewer strong options.",
        "resume":      "",
        "cover":       "",
        "fill":        YELLOW,
    },
    {
        "job_id":      "JOB-022",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Nagarro",
        "role":        "Associate Staff Engineer, Generative AI",
        "location":    "India (Chennai/Pune/Gurugram/Remote)",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/nagarro/jobs/associate-staff-engineer-generative-ai",
        "score":       85,
        "rationale":   "Near-perfect match: LangChain, LlamaIndex, RAG pipeline design, and Agentic AI frameworks "
                        "map directly onto GCP-GraphRAG and SAHAYAK-AI; role also wants requirement-to-design "
                        "translation and architecture documentation, both direct C-DAC experience. India remote, "
                        "large stable employer (5000+ employees).",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Top pick this run (score 85). Requires only 4+ yrs total experience vs 12+ yrs profile — "
                        "comfortably exceeds. Multiple locations posted (Chennai/Pune/Gurugram/Remote) — cover "
                        "letter written generically for remote/India.",
        "resume":      "jobs/Nagarro_AssociateStaffEngineer-GenerativeAI/Nagarro_AssociateStaffEngineer-GenerativeAI_resume.docx",
        "cover":       "jobs/Nagarro_AssociateStaffEngineer-GenerativeAI/Nagarro_AssociateStaffEngineer-GenerativeAI_cover.docx",
        "fill":        GREEN,
    },
    {
        "job_id":      "JOB-023",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Uvation",
        "role":        "AI/ML Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/uvation/jobs/ai-ml-engineer",
        "score":       58,
        "rationale":   "Generic ML engineering role (classification/regression/CV/time-series, MLOps with "
                        "Docker/Flask/Kubeflow) with no GenAI/LLM/RAG focus — profile's core strength (GenAI "
                        "research, RAG, agentic AI) is largely unused here.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 58 — below 70 threshold. Not tailored. Has an AI/ML component so not a deal-breaker, "
                        "just a weak fit relative to profile's GenAI specialisation.",
        "resume":      "",
        "cover":       "",
        "fill":        PINK,
    },
    {
        "job_id":      "JOB-024",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Pythian",
        "role":        "Senior AI / ML Engineer",
        "location":    "India",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/pythian/jobs/senior-ai-ml-engineer",
        "score":       70,
        "rationale":   "Solid applied AI/ML engineering role deploying LLMs and custom models to production "
                        "across AWS/GCP/Azure with Docker/Kubernetes — meets threshold via LLM-deployment "
                        "component, though less research/RAG-heavy than profile's GraphRAG focus.",
        "status":      "Scored",
        "dealbreaker": "No",
        "notes":       "Score 70 — exactly at threshold, not top 2 this run. India remote consultancy (201-500 "
                        "employees). Could revisit as a backup if a top-2 candidate falls through.",
        "resume":      "",
        "cover":       "",
        "fill":        YELLOW,
    },
    {
        "job_id":      "JOB-025",
        "date":        TODAY,
        "source":      "Web Search (Himalayas.app)",
        "company":     "Cotiviti",
        "role":        "Principal AI Engineer",
        "location":    "India / Remote (global)",
        "work_mode":   "Remote",
        "url":         "https://himalayas.app/companies/cotiviti/jobs/principal-ai-engineer",
        "score":       83,
        "rationale":   "Building a centralized agentic framework/AI platform directly parallels SAHAYAK-AI's "
                        "multi-agent MCP architecture; healthcare-payment-accuracy domain matches HMIS/healthcare-AI "
                        "delivery background. Principal level fits 12+ yrs seniority.",
        "status":      "Tailored",
        "dealbreaker": "No",
        "notes":       "Second top pick (score 83). Healthcare-tech company (5000+ employees) hiring remote across "
                        "India, Canada, Australia, Philippines, Mexico. Gap: payment/claims-specific fintech domain "
                        "not explicit in profile — cover letter leans on HMIS healthcare delivery instead.",
        "resume":      "jobs/Cotiviti_PrincipalAIEngineer/Cotiviti_PrincipalAIEngineer_resume.docx",
        "cover":       "jobs/Cotiviti_PrincipalAIEngineer/Cotiviti_PrincipalAIEngineer_cover.docx",
        "fill":        GREEN,
    },
]

wb = load_workbook(WB_PATH)
ws = wb['Applications'] if 'Applications' in wb.sheetnames else wb.active

for job in jobs:
    row = [
        job["job_id"],      # A: Job ID
        job["date"],        # B: Date sourced
        job["source"],      # C: Source
        job["company"],     # D: Company
        job["role"],        # E: Role
        job["location"],    # F: Location
        job["work_mode"],   # G: Work mode
        job["url"],         # H: Job URL
        job["score"],       # I: Fit score
        job["rationale"],   # J: Score rationale
        job["status"],      # K: Status
        job["dealbreaker"], # L: Deal-breaker?
        job["notes"],       # M: Missing requirements / notes
        job["resume"],      # N: Resume file
        job["cover"],       # O: Cover letter file
        "",                 # P: Date submitted
        "",                 # Q: Follow-up date
        "",                 # R: Outcome
        TODAY,              # S: Last updated
    ]

    ws.append(row)
    last_row = ws.max_row

    for col_idx in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.fill = job["fill"]
        cell.alignment = WRAP
        if col_idx == 1:
            cell.font = BOLD

wb.save(WB_PATH)
print(f"Tracker updated: JOB-021 through JOB-025 appended to {WB_PATH}")
